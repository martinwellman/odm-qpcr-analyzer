"""
# custom_functions.py

All available Excel custom functions in the Excel template files. Custom functions all begin with two underscores (\__). The return value 
of the custom function is used to replace the custom function call in the cell's value.
"""

import re
from openpyxl.utils import get_column_letter
from openpyxl.formula.translate import Translator
from openpyxl.utils.cell import coordinate_from_string
import math
import numbers

from qpcr_utils import (
    add_sheet_name_to_colrow_name,
)

# Regex to search for custom functions. %s should be the function name, eg __GETRANGE
# $1 of the result is the character immediately preceding the custom function. This should be a non-alphanumeric character. 
# It can also be CUSTOM_FUNC_SEP (;). If it is CUSTOM_FUNC_SEP, then when we replace the custom function with another string we also
# replace the CUSTOM_FUNC_SEP. eg. '=33;__SETCELL("Cal-abcd", "intercept", 44)' would result in something like '=3344'
# $2 is the custom function (eg. __GETRANGE)
# $3 is the function parameter list as a comma-separated string, (eg. '"Cal-abc", "intercept", 44')
CUSTOM_FUNCS_REGEX = "([^A-Za-z0-9_]|^)?(%s)\(([^\)]*)\)"

CUSTOM_FUNC_SEP = ";"

CUSTOM_FUNCS = {
    "__GETRANGE" : {
        "func" : "custom_func_getrange",
        "bind" : -10,
    },
    "__SETCELL" : {
        "func" : "custom_func_setcell",
        "bind" : 1,
    },
    "__QUOTIFY" : {
        "func" : "custom_func_quotify",
        "bind" : -20,
    },
    "__UNQUOTIFY" : {
        "func" : "custom_func_unquotify",
        "bind" : -20,
    },
    "__UPPER" : {
        "func" : "custom_func_upper",
        "bind" : -21,
    },
    "__LOWER" : {
        "func" : "custom_func_lower",
        "bind" : -21,
    },
    "__MAKEHEADER" : {
        "func" : "custom_func_makeheader",
        "bind" : -21,
    },
    "__SELECT" : {
        "func" : "custom_func_select",
        "bind" : -21,
    },
    "__GETDATA" : {
        "func" : "custom_func_getdata",
        "bind" : -1,
    },
    "__ADDROWID" : {
        "func" : "custom_func_addrowid",
        "bind" : 1,
    },
    "__GETCELL" : {
        "func" : "custom_func_getcell",
        "bind" : -10,
    },
    "__MERGETO" : {
        "func" : "custom_func_mergeto",
        "bind" : 20,
    },
    "__AVERAGE" : {
        "func" : "custom_func_average",
        "bind" : 0,
    },
    "__MOVINGAVERAGE" : {
        "func" : "custom_func_movingaverage",
        "bind" : 1,
    },
    "__GETCALVAL" : {
        "func" : "custom_func_getcalval",
        "bind" : -2,
    },
    "__QAQCHASFAILEDCATEGORY" : {
        "func" : "custom_func_qaqchasfailedcategory",
        "bind" : 1,
    }
}

def cast_bool(value):
    """Cast a value to a bool. Will properly detect strings such as "True", "T", "1".
    """
    if isinstance(value, str):
        return value.strip().lower() in ["t", "true", "1"]
    return bool(value)

def cast_int(value):
    """Cast a value to an int, or None if it can't be cast.
    """
    try:
        value = int(value)
    except:
        value = None
    return value

def custom_func_qaqchasfailedcategory(populator, category, failed, success, empty, target_sheet_name=None, target_cell=None, **kwargs):
    """Get an Excel formula that checks if the specified QAQC category has any failed values, and outputs a message to the user.

    Parameters
    ----------
    category : str
        The category name. These are the category values in the qaqc.yaml file, and are in the QAQC output sheets under
        the column "Category".
    failed : str
        The string to show if there is a failed QAQC check for the category.
    success : str
        The string to show if all QAQC checks for the category are good.
    empty : str
        The string to show if there are no QAQC checks for the category (which is typically interpreted as being good)
    target_sheet_name : str
        The sheet name where the formula will be found in.
    target_cell : openpyxl.cell.cell.Cell
        The cell that will contain the formula.
    """
    if populator.hide_qaqc or populator.qaqc == None:
        return '""'

    ws, _ = populator.get_worksheet_and_info(target_sheet_name)
    return populator.qaqc.add_failed_category_qaqc_cell(category, failed, success, empty, ws, target_cell.coordinate)

def custom_func_getcalval(populator, sheet_name, cal_value, **kwargs):
    """Get the specified calibration curve value for the curve with ID sheet_name.

    Parameters
    ----------
    sheet_name : str
        The name of the calibration curve.
    cal_value : str
        The ID of the value to retrieve. This is any value in the calibration curve's sheet info object
        (found in worksheets_info). See get_calibration_value for all possible values.
    """
    value = populator.get_calibration_value(sheet_name, cal_value)
    if value is None:
        raise ValueError(f"ERROR: Cal curve '{sheet_name}' not found")
    return value

def custom_func_getrange(populator, row_name, col_name, max_items=None, template_cell=None, **kwargs):
    """
    Get an Excel range formula for all cells with the specified row_name and col_name. These are the names specified in the template.

    Parameters
    ----------
    row_name : str
        The row name to retrieve the range for (eg. "main_row_data").
    col_name : str
        The column name to retrieve the range for (eg. "main_col_ct").
    max_items : int
        Include at most this many rows or columns.
    target_sheet_name : str
        The name of the sheet to get the range in.
    """
    fixed_rows = False #cast_bool(fixed_rows)
    fixed_cols = False #cast_bool(fixed_cols)
    max_items = cast_int(max_items)
    target_sheet_name = kwargs.get("target_sheet_name", None)
    if max_items is not None:
        max_items = int(max_items)
    return populator.get_named_range(target_sheet_name, row_name, col_name, fixed_rows=fixed_rows, fixed_cols=fixed_cols, max_rows=max_items, max_cols=max_items)

def custom_func_getcell(populator, sheet_name, id, prefer_precalculated=True, default="\"{id} missing\"", **kwargs):
    """Get the cell address (or literal value) for the cell with the specified ID in the specified sheet.

    Parameters
    ----------
    sheet_name : str
        The sheet name to get the cell from.
    id : str
        The ID of the cell to get. This is either an ID set by a call to __SETCELL or the id of a calibration curve value.
    prefer_precalculated : bool
        If True, and we have the literal value available for the cell, then retrieve the precalculated value instead.
        This is for all calibration curve values. If the precalculated value does not exist then the
        cell address will be retrieved instead.
    default : str
        If the cell or value does not exist then return this value. The tags {sheet_name} and {id} are replaced with the corresponding
        parameters.
    target_sheet_name : str
        The sheet name that will receive the value (ie. will contain the cell reference). This is used to
        determine if the cell address requires its sheet name (eg. 'Main'!A5)
    """
    fixed_row = True #cast_bool(fixed_row)
    fixed_col = True #cast_bool(fixed_col)
    prefer_precalculated = cast_bool(prefer_precalculated)
    target_sheet_name = kwargs.get("target_sheet_name", None)
    addr = populator.get_named_cell_address_or_value(sheet_name, id, fixed_row=fixed_row, fixed_col=fixed_col, prefer_precalculated=prefer_precalculated, target_sheet_name=target_sheet_name)

    if addr is None or (isinstance(addr, numbers.Number) and math.isnan(addr)):
        if default is None:
            raise ValueError(f"ERROR: Cell '{id}' in '{sheet_name}' not found")
        else:
            if isinstance(default, str):
                default = default.format(id=id, sheet_name=sheet_name)
            return default

    return addr

def custom_func_addrowid(populator, sheet_name, id, replace_value="", **kwargs):
    """Add the specified ID to the current row. A row ID is the same as a row name, it is added
    to the worksheets_info object for the sheet, to the "row_names" member.

    Parameters
    ----------
    sheet_name : str
        The sheet name to add the row name to.
    id : str    
        The row name to add to the current row.
    replace_value : str
        The value to replace the function call (__ADDROWID) with. This will be the return value of this function.
    """
    target_cell = kwargs.get("target_cell", None)
    target_sheet_name = kwargs.get("target_sheet_name", None)
    _, info = populator.get_worksheet_and_info(sheet_name)
    if info is not None:
        cur_row = target_cell.row - info["origin"][0]
        name = add_sheet_name_to_colrow_name(target_sheet_name, id)
        info["row_names"][cur_row].append(name)

    return replace_value

def custom_func_movingaverage(populator, days, center_cell, date_cell, match_cell, template_cell, target_cell, **kwargs):
    """Get an Excel formula that will calculate the moving average centered on a cell.

    Parameters
    ----------
    days : int | str
        The number of days for the moving average.
    center_cell : str
        The cell address for the center of the moving average. We will get the cell addresses in a vertical block of size
        days surrounding the center_cell.
    date_cell : str
        The cell address specifying the date of the current cell. We will use this cell and address to calculate the
        day range of the moving average.
    match_cell : str,
        The cell address specifying the id to match for inclusion into the moving average calculation. All cells that
        have the same value at match_cell in the column for match_cell are included (provided it is in the 5-day
        range). eg. This can be a Site ID, such as "O", in which case all rows for site ID "O" are potentially
        included in the average calculation.
    template_cell : openpyxl.cell.cell.Cell
        The cell that is being used as the template cell for the current cell. This is required because we need to translate
        the center_cell address from the template_cell to the target_cell.
    target_cell : openpyxl.cell.cell.Cell
        The cell that is making the function call. It will receive the formula.
    """
    center_cell = Translator(f"={center_cell}", template_cell.coordinate).translate_formula(target_cell.coordinate)
    center_cell = center_cell[1:]
    date_cell = Translator(f"={date_cell}", template_cell.coordinate).translate_formula(target_cell.coordinate)
    date_cell = date_cell[1:]
    match_cell = Translator(f"={match_cell}", template_cell.coordinate).translate_formula(target_cell.coordinate)
    match_cell = match_cell[1:]

    days = int(days)

    center_column, center_row = coordinate_from_string(center_cell)
    match_column, match_row = coordinate_from_string(match_cell)
    date_column, date_row = coordinate_from_string(date_cell)
    delta_ahead = days // 2
    delta_behind = days - delta_ahead - 1
    # =AVERAGEIFS(D:D,B:B,"="&B2,C:C,"<="&(C2+2),C:C,">="&(C2-2))
    formula = f'AVERAGEIFS({center_column}:{center_column},{match_column}:{match_column},"="&{match_cell},{date_column}:{date_column},"<="&({date_cell}+{delta_ahead}), {date_column}:{date_column},">="&({date_cell}-{delta_behind}))'
    # rng = f'INDIRECT(ADDRESS(ROW({center_cell})-{delta_up},COLUMN({center_cell}))&":"&ADDRESS(ROW({center_cell})+{delta_down},COLUMN({center_cell})))'
    # formula = f'IF(ROW()>={delta_up}+1,IF(COUNT({rng})={days},AVERAGE({rng}),""),"")'
    return formula

def custom_func_average(populator, *args, **kwargs):
    """Get an Excel formula to calculate the average of multiple values.

    Parameters
    ----------
    args : list
        A list of all values to calcualte the average of. Any value that is empty ("") or nan ("nan") are ignored.
    """
    args = [a for a in args if a not in ["", "nan"]]
    if len(args) == 0:
        return "" ##DIV/0!"

    return "AVERAGE({})".format(",".join(args))

def custom_func_mergeto(populator, location, replace_value="", **kwargs):
    """Merge the current Excel cell in a certain direction.

    Parameters
    ----------
    location : str
        The direction to merge to. Can be:
            "bottom" : Merge to the bottom extent of the current analysis group.
            "right" : Merge to the right extent of the current analysis group.
            "top" : Merge to the top extent (origin) of the current analysis group.
            "left" : Merge to the left extent (origin) of the current analysis group.
    replace_value : str
        The value to replace the __MERGETO function call with, in the cell.
    """
    target_cell = kwargs.get("target_cell", None)
    target_sheet_name = kwargs.get("target_sheet_name", None)
    ws, info = populator.get_worksheet_and_info(target_sheet_name)

    def _copy_to_origin_cell(ws, row_a, col_a, row_b, col_b):
        orig_cell = ws[f"{col_a}{row_a}"]
        source_cell = ws[f"{col_b}{row_b}"]
        if isinstance(source_cell.value, str):
            source_cell.value = re.sub(CUSTOM_FUNCS_REGEX % "__MERGETO", "$1")
        populator.copy_cell(source_cell, orig_cell)
    
    col_a = get_column_letter(target_cell.column)
    row_a = target_cell.row
    if location == "bottom":
        row_b = info["extents"][0]
        row_b = max(row_b-1, 1)
        col_b = col_a
    elif location == "right":
        row_b = row_a
        col_b = get_column_letter(info["extents"][1]-1)
    elif location == "top":
        row_b = row_a
        row_a = info["origin"][0]
        col_b = col_a
        _copy_to_origin_cell(ws, row_a, col_a, row_b, col_b)
    elif location == "left":
        col_b = col_a
        row_b = row_a
        col_a = get_column_letter(info["origin"][1])
        _copy_to_origin_cell(ws, row_a, col_a, row_b, col_b)

    merge_range = f"{col_a}{row_a}:{col_b}{row_b}"
    ws.merge_cells(merge_range)

    return replace_value

def custom_func_setcell(populator, sheet_name, id, replace_value="", **kwargs):
    """Set an ID for teh current cell, so it can be retrieved later with a call to __GETCELL

    Parameters
    ----------
    sheet_name : str
        The sheet name to set the cell in. This is the "sheet_name" in the worksheets_info object.
    id : str
        The ID to give to the cell.
    replace_value : str
        The value to replace the function call with (ie. the return value).
    """
    target_cell = kwargs.get("target_cell", None)
    _, info = populator.get_worksheet_and_info(sheet_name)
    if info is not None:
        info[id] = target_cell.coordinate

    return replace_value

def custom_func_select(populator, value_a, value_b, **kwargs):
    """Select value_a if it is not blank, otherwise select value_b.
    """
    if value_a is None or value_a == "":
        return value_b
    return value_a

def custom_func_makeheader(populator, value, **kwargs):
    """Convert text to header format: Replace non-alphanumeric characters with underscore and make lowercase.
    """
    if isinstance(value, str):
        value = re.sub("[^A-Za-z0-9]", "_", value).lower()
    return value
    

def custom_func_upper(populator, value, **kwargs):
    """Make text uppercase.
    """
    if isinstance(value, str):
        value = value.upper()
    return value

def custom_func_lower(populator, value, **kwargs):
    """Make text lowerrcase.
    """
    if isinstance(value, str):
        value = value.lower()
    return value

def custom_func_quotify(populator, value, **kwargs):
    """Add quotes to the value.
    """
    value = value if value is None else str(value)
    value = value.replace('"', '""')
    return f'"{value}"'

def custom_func_unquotify(populator, value, only_if_number=False, **kwargs):
    """Remove quotes from the value.
    """
    number_value = value if value is None else str(value)
    number_value = number_value.replace('"', '').replace("'", "")
    if cast_bool(only_if_number):
        try:
            number_value = float(number_value)
        except:
            return f'"{value}"'
    return number_value

def custom_func_getdata(populator, cell_addr, id, format=None, **kwargs):
    """Get the values in the attached_data of the cell.

    Parameters
    ----------
    cell_addr : str
        The cell address to get the data of.
    id : str
        The column(s) to get the values from. We find the first attached_data that has each of the columns
        and get the value from it. To retrieve multiple columns separate the columns with a semi-colon.
    format : str
        An optional string format tag. If None then we return all values in id, joined with semi-colons.
        If not null, then we return format.format(*all_values). ie. {0} is replaced with the value
        found for id[0], {1} for id[1], etc.
    """
    target_sheet_name = kwargs.get("target_sheet_name", None)
    ws, info = populator.get_worksheet_and_info(target_sheet_name)
    cell = ws[cell_addr]
    # attached_data = populator.get_cell_attached_data(cell)

    all_ids = [x.strip() for x in id.split(";")]
    all_values = []
    for idx, cur_id in enumerate(all_ids):
        val = populator.get_cell_attached_data_value(cell, cur_id, "")
        all_values.append(val)
    
    if format is None:
        return ";".join([str(a) for a in all_values])

    return format.format(*all_values)
