#%%
# %load_ext autoreload
# %autoreload 2

"""
# qpcr_updater.py

Updates master QPCR data files with new and approved data, by appending new rows to the end of the master Excel file.

The input files should be in a format with the first row being the column names, and each subsequent row being a data row
that should be copied to the master file.

The rows are copied in the exact column order as the input file. We do not check the master file to make sure the column names match
or are in the same order. We also only update the master file if the input file has the columns specified in required_source_headers
in the config file.

The sheet that is copied to in the target master workbook is named main_worksheet (in the config file).

In order to determine the target master workbook, we check for the site ID in column site_column (in the config file) in the
input file.

## Usage

    updater = QPCRUpdater(updater_config, 
        populator_config, 
        sites_config=sites_config, 
        sites_file=sites_file)
    # updated is a list of bools. If True at index n, then input_files[n] has been appended to the target_path_template.
    # If False at index n, then input_files[n] was not appended to target_path_template, due to not being in the correct format.
    # target_path_template can contain tags specific to the site ID: {site_id}, {site_title}, {parent_site_id}, {parent_site_title},
    # and {sample_type}
    updated = updater.update(input_files, target_path_template)
"""

import os
from easydict import EasyDict
import argparse
from datetime import datetime
import openpyxl
import yaml
from openpyxl.utils import get_column_letter
from openpyxl.formula.translate import Translator
import cloud_utils
import gdrive_utils
from qpcr_utils import (
    load_config,
)
from fix_xlsx import fix_xlsx_file
from excel_calculator import add_excel_calculated_values
import tempfile
from copy import copy
from qpcr_sites import QPCRSites

import logging
logging.getLogger("pycel").setLevel(logging.CRITICAL)

class QPCRUpdater(object):
    def __init__(self, config_file, populator_config_file, sites_config, sites_file):
        super().__init__()
        self.local_dir = tempfile.gettempdir()
        self.target_workbooks = {}

        self.config = load_config(config_file)
        self.populator_config = load_config(populator_config_file)
        self.sites = QPCRSites(config_file=sites_config, sites_file=sites_file)

    def get_columns(self, worksheet):
        """Get all column names from the worksheet. These are the values found in row 1, and are the columns we'd
        normally get when loading with Pandas.
        """
        max_col = worksheet.max_column
        cells = worksheet[f"A1:{get_column_letter(max_col)}1"]
        columns = [c.value for c in cells[0]]
        return columns

    def get_workbook_info(self, site_id):
        """Get the info of the workbook (ie. Excel file) for the specified site_id. The info is a dictionary with multiple
        values associated with the workbook, including the remote_target (target file to add to), the local_target (local 
        file to add to), and the actual "wb" (workbook).
        """
        remote_target = self.get_remote_target_file(site_id)
        return self.target_workbooks.get(remote_target, None)

    def get_remote_target_file(self, site_id):
        return self.sites.parse_filename_for_siteid(self.target_path_template, site_id)

    def remove_trailing_blank_lines(self, ws):
        """Remove all lines at the bottom of the worksheet that are blank.
        """
        max_row = ws.max_row
        max_column = get_column_letter(ws.max_column)
        top_blank_row = None
        # Find the first blank line at end of worksheet (top_blank_row)
        for row_num in range(max_row, 0, -1):
            row = ws[f"A{row_num}:{max_column}{row_num}"]
            row_values = [cell.value for cell in row[0] if cell.value]
            if len(row_values) > 0:
                break
            top_blank_row = row_num
        if top_blank_row is not None:
            print(f"Deleting {max_row-top_blank_row+1} trailing blank rows, starting at row {top_blank_row}")
            ws.delete_rows(top_blank_row, max_row - top_blank_row + 1)

    def get_target_workbook(self, site_id):
        """Get the target OpenPYXL workbook that we copy to for the specified site_id.
        """
        site_info = self.get_workbook_info(site_id)
        if site_info is not None:
            # The workbook already exists
            return site_info["wb"]
        else:
            print(f"Downloading site workbook for {site_id}")
            remote_target = self.get_remote_target_file(site_id)
            local_target = cloud_utils.download_file(remote_target)

            if local_target and os.path.isfile(local_target):
                # File was downloaded so load it
                fix_xlsx_file(local_target)
                wb = openpyxl.load_workbook(local_target)
                self.remove_trailing_blank_lines(self.get_main_ws(wb))
            else:
                # File was not downloaded so create a new workbook
                local_target = os.path.join(self.local_dir, cloud_utils.remove_prefix(remote_target))
                wb = openpyxl.Workbook()

            self.target_workbooks[remote_target] = {
                "local_target" : local_target,
                "wb" : wb,
            }

            return wb
            
    def get_main_ws(self, wb):
        if self.config.main_worksheet in wb.sheetnames:
            return wb[self.config.main_worksheet]
        else:
            return wb[wb.sheetnames[0]]

    def copy_cell(self, source_cell, target_cell):
        """Copy the source cell to the target cell, including style options.
        """
        if isinstance(source_cell.value, str):
            target_cell.value = Translator(source_cell.value, source_cell.coordinate).translate_formula(target_cell.coordinate)
        else:
            target_cell.value = source_cell.value
        if source_cell.has_style:
            # styles = [s for s in self.output_wb._named_styles if s.name == source_cell.style]
            # if len(styles) > 0:
            #     target_cell.style = styles[0]
            target_cell.fill = copy(source_cell.fill)
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)
        return target_cell

    def update(self, input_files, target_path_template):
        """Do a full update by copying valid input_files to the target files.

        Parameters
        ----------
        input_files : list[str]
            List of local Excel files that we want to copy to the target files.
        target_path_template : str
            The target file name, including the path. Can be a remote target (ie. preceded with gd:// for
            file on Google Drive or s3:// for file on AWS S3) or a local file. Tags that can be used are
            {site_id}, {site_title}, {parent_site_id}, {parent_site_title}, and {sample_type}. The site ID is determined
            by extracting it from the input file under the config.site_column column.

        Returns
        -------
        list[bool]
            A list of bools representing if input_files[n] was copied to the target file or not. If it was copied
            then all the required_source_headers columns in the input file were present. If it was not copied
            it was because those columns were not available in the input file.
        """
        self.target_path_template = target_path_template
        success = []

        for input_file in input_files:
            try:
                fix_xlsx_file(input_file)
                wb = openpyxl.load_workbook(input_file)
            except:
                success.append(False)
                continue
            # add_excel_calculated_values(wb)
            ws = wb[wb.sheetnames[0]]
            columns = [(c or "").strip().lower() for c in self.get_columns(ws)]

            # Make sure all of the required columns are present in the input file
            found_columns = [col for col in self.config.required_source_headers if col.strip().lower() in columns]
            if len(found_columns) < len(self.config.required_source_headers):
                missing_columns = list(set(self.config.required_source_headers) - set(found_columns))
                print(f"Required column(s) {missing_columns} not found in input file {input_file}")
                success.append(False)
                continue

            site_col = columns.index(self.config.site_column.strip().lower())

            # Go through each row in the input file, copy them to the target file matching the site ID for that row
            for row in ws.iter_rows(min_row=2):
                if self.is_empty_row(row):
                    continue

                site_id = row[site_col].value
                target_wb = self.get_target_workbook(site_id)
                target_ws = self.get_main_ws(target_wb)

                target_wb.loaded_theme = wb.loaded_theme

                # If worksheet is empty then add the headings
                if target_ws.max_row <= 1 and target_ws.max_column <= 1:
                    for idx, header_cell in enumerate(ws["1:1"]):
                        col = get_column_letter(idx+1)
                        self.copy_cell(header_cell, target_ws[f"{get_column_letter(idx+1)}1"])
                        target_ws.column_dimensions[col].width = ws.column_dimensions[col].width
                    # Merge the headings
                    for rng in ws.merged_cells.ranges:
                        min_row, min_col = rng.min_row, rng.min_col
                        max_row, max_col = rng.max_row, rng.max_col
                        if min_row == max_row and min_row == 1:
                            merge_range = "{}{}:{}{}".format(get_column_letter(min_col), min_row, get_column_letter(max_col), max_row)
                            target_ws.merge_cells(merge_range)
                    if ws.row_dimensions[1].height:
                        target_ws.row_dimensions[1].height = ws.row_dimensions[1].height

                # Add the row
                cur_row = target_ws.max_row+1
                for col,cell in enumerate(row):                    
                    self.copy_cell(cell, target_ws[f"{get_column_letter(col+1)}{cur_row}"])

            success.append(True)
        
        # remote_targets = self.save_all()

        return success
    
    def is_empty_row(self, row):
        for cell in row:
            if cell.value is not None and len(str(cell.value)) > 0:
                return False
        return True

    def save_and_upload(self):
        """Save all our target output files to disk, and upload them to the remote target if the
        target_path_template sent to update is a remote path.

        Returns
        -------
        list[str]
            List of where all the output target files were uploaded to. These can be remote targets (preceded by
            s3:// or gd://) or local paths, depending on the target_dir passed in the previous call to update.
        """
        remote_targets = []
        # for wb_info in self.target_workbooks:
        for remote_target, wb_info in self.target_workbooks.items():
            local_target = wb_info["local_target"]
            local_dir = os.path.dirname(local_target)
            if local_dir:
                os.makedirs(local_dir, exist_ok=True)
            add_excel_calculated_values(wb_info["wb"])
            print(f"Saving to {local_target}")
            wb_info["wb"].save(local_target)
            print(f"Uploading to {remote_target}")
            cloud_utils.upload_file(local_target, remote_target)
            remote_targets.append(remote_target)
        return remote_targets

if __name__ == "__main__":
    if "get_ipython" in globals():
        opts = EasyDict({
            "input_files" : [
                "/Users/martinwellman/Documents/Health/Wastewater/Code/inputs/biorad-ottawa/sep10/Data - <unk>.xlsx",
                # "/Users/martinwellman/Documents/Health/Wastewater/Code/temp.xlsx",
                # "/Users/martinwellman/Documents/Health/Wastewater/Code/populated-wide/Data - Uottawa.xlsx",
                # "/Users/martinwellman/Documents/Health/Wastewater/Code/populated-wide/Data - Ottawa.xlsx",
                # "/Users/martinwellman/Documents/Health/Wastewater/Code/populated-wide/Data - Nippising (First Nations).xlsx",
            ],
            # "target_dir" : "gd://{parent_site_title}/",
            # "target_dir" : "gd://hidden_test/",
            "target_dir" : "/Users/martinwellman/Documents/Health/Wastewater/Code/test-new/new/",
            # "target_path_template" : "Data - {parent_site_title}.xlsx",
            "target_path_template" : "Data - All - {parent_site_title}.xlsx",
            "config" : "qpcr_updater.yaml",
            "populator_config" : "qpcr_populator_ottawa.yaml",
            "sites_config" : "sites.yaml",
            "sites_file" : "sites.xlsx",
        })
    else:
        args = argparse.ArgumentParser(formatter_class=argparse.ArgumentDefaultsHelpFormatter)
        args.add_argument("--config", nargs="+", type="str", help="Config file to use", required=True)
        args.add_argument("--populator_config", nargs="+", type="str", help="QPCRPopulator config file(s) used when the input_files were generated.", required=True)
        args.add_argument("--target_dir", type="str", help="Target folder (possibly remote) where the already populated outputs are. (We append to these)", required=True)
        args.add_argument("--target_path_template", type="str", help="Name template for the output files. Can contain tags {site_id}, {site_title}, {parent_site_id}, {parent_site_title}, and {sample_type}", required=True)
        args.add_argument("--input_files", nargs="+", type="str", help="Input files containing the data", required=True)
        args.add_argument("--sites_file", type=str, help="Excel file specifying all WW sites with information about each site.", required=False)
        args.add_argument("--sites_config", type=str, help="Config file for the sites file.", required=False)
        opts = args.parse_args()
    
    with open("../../event.json", "r") as f:
        data = EasyDict(yaml.safe_load(f))
    gdrive_utils.set_creds_file("credentials.json")
    gdrive_utils.set_partial_token_data(data.tokens)

    tic = datetime.now()
    updater = QPCRUpdater(opts.config, opts.populator_config, sites_config=opts.sites_config, sites_file=opts.sites_file)
    print("Update:", updater.update(opts.input_files, os.path.join(opts.target_dir, opts.target_path_template)))
    print("Upload:", updater.save_and_upload())
    toc = datetime.now()
    print("Total duration:", toc - tic)


