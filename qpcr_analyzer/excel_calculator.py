#%%
# %load_ext autoreload
# %autoreload 2
"""
excel_calculator.py
===================

See [excel_calculator.md](excel_calculator.md) for details.

Usage
-----

    add_excel_calculated_values(wb)
    wb.save("output_file.xlsx")

For some unknown reason, we sometimes need to first save `wb` to disk then reload it before calling `add_excel_calculated_value`.
Until this is fixed, it is recommended to always save and reload first:

    wb.save("output_file.xlsx")
    wb = openpyxl.load_workbook("output_file.xlsx")
    add_excel_calculated_values(wb)
    wb.save("output_file.xlsx")

"""

from pycel.excelutil import (
    ERROR_CODES,
    NA_ERROR,
    REF_ERROR,
    VALUE_ERROR,
)

from openpyxl.cell.cell import Cell
import inspect
from pycel.excelformula import FormulaEvalError
from openpyxl.utils import get_column_letter
import openpyxl

if 'calculated_value' not in Cell.__slots__:
    path = inspect.getfile(Cell)
    raise RuntimeError(f"OpenPYXL Cell class requires custom 'calculated_value' in its __slots__, Cell is defined in: {path}")

import pycel
from pycel import ExcelCompiler
from pycel.lib.function_helpers import (
    excel_helper,
)

import pycel.lib.stats
import pycel.lib.lookup
from pycel.excellib import _numerics
from pycel.excelutil import (
    DIV0,
)
import numpy as np

def add_excel_calculated_values(xl, sheets=None):
    """Add calculated values to all workbook cells in the specified sheets (list of sheet names). If sheets is None then all sheets are
    calculated.
    """
    excel = ExcelCompiler(excel=xl)

    sheets = xl.sheetnames if sheets is None else sheets
    for sheet_name in sheets:
        print(f"Adding calculated values to {sheet_name}")
        sheet = xl[sheet_name]
        for row_num in range(sheet.min_row, sheet.max_row+1):
            for cell in sheet[row_num]:
                if cell.data_type == "f" and isinstance(cell.value, str) and cell.value.strip()[0:1] == '=':
                    addr = f"'{sheet_name}'!{cell.coordinate}"

                    try:
                        val = excel.evaluate(addr)
                        cell.calculated_value = val
                    except NameError:
                        cell.calculated_value = "#NAME?"
                    except FormulaEvalError as e:
                        pass
                    except Exception as e:
                        raise ValueError(f"ERROR evaluating Excel formula at {addr}: {cell.value}: {e}")

@excel_helper()
def stdev(*args):
    # Excel reference: https://support.microsoft.com/en-us/office/
    #   average-function-047bac88-d466-426c-a32b-8f33eb960cf6
    data = _numerics(*args)

    # A returned string is an error code
    if isinstance(data, str):
        return data
    elif len(data) == 0:
        return DIV0
    else:
        return np.std(data, ddof=1)

def _slope_intercept(Y, X, **kwargs):
    XY = [[x[0],y[0]] for x,y in zip(X,Y) if isinstance(x[0], (float, int)) and isinstance(y[0], (float, int))]
    Y = tuple([(y,) for _,y in XY])
    X = tuple([(x,) for x,_ in XY])

    try:
        coefs, full_rank = pycel.lib.stats.linest_helper(Y, X, **kwargs)
    except AssertionError:
        return NA_ERROR
    except ValueError:
        return VALUE_ERROR

    if len(coefs) < 2:
        return NA_ERROR
    if not full_rank:
        return DIV0
    return coefs


@excel_helper()
def slope(Y, X):
    # Excel reference: https://support.microsoft.com/en-us/office/
    #   slope-function-11fb8f97-3117-4813-98aa-61d7e01276b9
    coefs = _slope_intercept(Y, X)
    if coefs in ERROR_CODES:
        return coefs
    return coefs[0]

@excel_helper()
def rsq(Y, X):
    # Excel reference: https://support.microsoft.com/en-us/office/
    #   rsq-function-d7161715-250d-4a01-b80d-a8364f2be08f
    
    coefs = _slope_intercept(Y, X, stats=True)
    if coefs in ERROR_CODES:
        return coefs
    return coefs[2][0]


    # XY = [[x[0],y[0]] for x,y in zip(X,Y) if isinstance(x[0], (float, int)) and isinstance(y[0], (float, int))]
    # Y = tuple([(y,) for _,y in XY])
    # X = tuple([(x,) for x,_ in XY])

    # try:
    #     coefs, full_rank = pycel.lib.stats.linest_helper(Y, X, stats=True)
    # except AssertionError:
    #     return REF_ERROR
    # except ValueError:
    #     return VALUE_ERROR

    # return coefs[2][0]
    
@excel_helper()
def intercept(Y, X):
    # Excel reference: https://support.microsoft.com/en-us/office/
    #   intercept-function-2a9b74e2-9d47-4772-b663-3bca70bf63ef
    coefs = _slope_intercept(Y, X)
    if coefs in ERROR_CODES:
        return coefs
    return coefs[1]
    
@excel_helper()
def hyperlink(value, text):
    # Excel reference: https://support.microsoft.com/en-us/office/
    #   hyperlink-function-333c7ce6-c5ae-4164-9c47-7de9b76f577f
    return text

@excel_helper(ref_params=0)
def address(row, col):
    addr = "${}${}".format(get_column_letter(col), row)
    return addr

pycel.lib.stats.stdev = stdev
pycel.lib.stats.rsq = rsq
pycel.lib.stats.slope = slope
pycel.lib.stats.intercept = intercept
pycel.lib.lookup.hyperlink = hyperlink
pycel.lib.lookup.address = address

if __name__ == "__main__":
    wb = openpyxl.load_workbook("/Users/martinwellman/Documents/Health/Wastewater/Code/test-new/populated-sep10.xlsx")
    add_excel_calculated_values(wb)
    wb.save("/Users/martinwellman/Documents/Health/Wastewater/Code/test-new/calc.xlsx")
    print("Finished!")

