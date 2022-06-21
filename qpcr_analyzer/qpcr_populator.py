#%%
# %load_ext autoreload
# %autoreload 2

"""
# qpcr_populator.py

See [qpcr_populator.md](qpcr_populator.md).
"""

# This is required for the OpenPYXL changes made by localfixes.sh, since we only
# made the fixes for saving the XLSX file with the etree rather than LXML.
import os
os.environ["OPENPYXL_LXML"] = "False"

from easydict import EasyDict
from copy import copy, deepcopy
import pandas as pd
import numpy as np
import os
import re
import math
import cloud_utils
import argparse

from datetime import datetime, date

from excel_calculator import add_excel_calculated_values
import logging
logging.getLogger("pycel").setLevel(logging.CRITICAL)

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils import units
from openpyxl import Workbook
from openpyxl.formula.translate import Translator
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.trendline import Trendline, TrendlineLabel
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.styles.numbers import is_date_format

from qpcr_qaqc import (
    QPCRQAQC,
)

# pd.set_option('mode.chained_assignment', "raise")

from qpcr_sites import QPCRSites
from qpcr_sampleids import QPCRSampleIDs
from qpcr_sampleslog import QPCRSamplesLog
from qpcr_methods import QPCRMethods
from excel_file_utils import fix_xlsx_file
from qpcr_utils import (
    OUTLIER_COL,
    MAIN_SHEET,
    SINGLE_CAL_SHEET,
    CAL_SHEET_FMT,
    MAIN_ROW_DATA,
    CAL_ROW_DATA,
    points_to_cm,
    estimated_cm_to_chars,
    add_sheet_name_to_colrow_name,
    add_sheet_name_to_colrow_names,
    get_template_colrow_names,
    flatten,
    load_config,
    strip_quotes,
    excel_addr_to_fixed,
    sheet_to_df,
    parse_values,
)
import custom_functions
from custom_functions import (
    CUSTOM_FUNCS,
    CUSTOM_FUNCS_REGEX,
    CUSTOM_FUNC_SEP,
)

class QPCRPopulator(object):
    def __init__(self, input_file, template_file, target_file, overwrite, config_file, qaqc_config_file, sites_config, sites_file, sampleids_config, sampleslog_config, sampleslog_file, methods_config, methods_file, hide_qaqc=False):
        super().__init__()
        self.hide_qaqc = hide_qaqc
        self.input_file = input_file
        self.template_file = template_file
        self.target_file = target_file
        # self.local_target_file = cloud_utils.download_file(target_file)
        self.overwrite = overwrite
        self.config_file = config_file if isinstance(config_file, (list, tuple, np.ndarray)) else [config_file]
        # self.qaqc_config_file = qaqc_config_file if qaqc_config_file is None or isinstance(qaqc_config_file, (list, tuple, np.ndarray)) else [qaqc_config_file]
        self.worksheets_info = []
        self.late_binders = []
        self.row_data = {}

        self.config = load_config(self.config_file)
        self.qaqc = QPCRQAQC(self, qaqc_config_file, config_file)

        self.sites = QPCRSites(sites_config, sites_file)
        self.sampleids = QPCRSampleIDs(sampleids_config, sites_config, sites_file)
        self.sampleslog = QPCRSamplesLog(sampleslog_config, sampleslog_file, sampleids_config, sites_config, sites_file)
        self.methods = QPCRMethods(methods_config, methods_file)

    def get_column_names(self, sheet_name, col_id):
        """Get all the names (eg. "main_col_ct") attached to the specified Excel column (eg. "AB") in the specified sheet.
        """
        col = column_index_from_string(col_id)
        ws, info = self.get_worksheet_and_info(sheet_name)
        _, min_col = info["origin"]
        return info["col_names"][col - min_col]

    def get_named_columns(self, sheet_name, match_col_name, flatten_names=False):
        """Get the Excel column Ids (eg. 'A', 'AB') of the named columns, as specified in a
        template file. (These are the named columns in the formatted output file).
        
        Parameters
        ----------
        sheet_name : str
            The sheet name.
        match_col_name : str | list | tuple | np.ndarray 
            The column names to get the column Ids for.
        flatten_names : bool
            If True, then flatten the results into a 1-D array. If False, then the
            returned array is 2D, with arr[n] being an array for the columns with
            name match_col_name[n].
        
        Returns
        -------
        list
            The matched column names, in the format according to the parameter flatten_names.
        """
        # Clean up requested columns. Column names are case-insensitive, so make all lower case.
        if isinstance(match_col_name, str):
            match_col_name = [match_col_name]
            flatten_names = True
        match_col_name = [c.strip().lower() for c in match_col_name]
        
        columns = []
        for _ in range(len(match_col_name)):
            columns.append([])

        # Loop through all columns in the sheet, see if the column has one of the requested column names
        ws, info = self.get_worksheet_and_info(sheet_name)
        min_row, min_col = info["origin"]
        col_names = info["col_names"]
        for col_num, c in enumerate(col_names):            
            matches = [cur is not None and cur.strip().lower() in match_col_name for cur in c]
            if np.any(matches):
                addr = get_column_letter(col_num + min_col)
                idx = np.argmax(matches)
                match_col = c[idx]
                idx = match_col_name.index(match_col.lower())
                columns[idx].append(addr)

        if flatten_names:
            columns = flatten(columns)
        return columns

    def get_named_range(self, sheet_name, row_name, col_name, fixed_rows=True, fixed_cols=True, include_sheet_name=False, override_row=None, override_col=None, max_rows=None, max_cols=None):
        """Get an Excel range (eg. "AB3:AB20") for the specified row and column names.

        Parameters
        ----------
        sheet_name : str
            The name of the sheet to get the range in.
        row_name : str | list[str]
            The row names to retrieve the range for (eg. ["main_row_data"]).
        col_name : str | list[str]
            The column names to retrieve the range for (eg. ["main_col_ct"]).
        fixed_rows : bool
            If True then return the range using fixed rows format (eg. "AB$3:AB$20").
        fixed_cols : bool
            If True then return the range using fixed columns format (eg. "$AB3:$AB20").
        include_sheet_name : bool
            If True then include the sheet name in the returned range (eg. 'Main'!AB3:AB20).
        override_row : int
            Optional 1-based row to use in the range, rather than the rows that match the row_name (eg. 10)
        override_col : int
            Optional 1-based column to use in the range, rather than the rows that match the col_name (eg. 10)
        max_rows : int
            Optional maximum number of rows to include in the range. We clip the actual range to this.
        max_rows : int
            Optional maximum number of columns to include in the range. We clip the actual range to this.

        Returns
        -------
        str
            The Excel range (eg. "AB3:AB20")
        """
        row_name = row_name or []
        col_name = col_name or []
        if isinstance(row_name, str):
            row_name = [row_name]
        if isinstance(col_name, str):
            col_name = [col_name]

        # Add the sheet_name to the column/row names, since that is how we store the names internally
        col_name = [add_sheet_name_to_colrow_name(sheet_name, c) for c in col_name]
        row_name = [add_sheet_name_to_colrow_name(sheet_name, r) for r in row_name]
        # col_name, row_name = add_sheet_name_to_colrow_names(sheet_name, col_name, row_name)

        target_ws, target_info = self.get_worksheet_and_info(sheet_name)

        # Get the current origin/extents, and the current column and row names of the sheet.
        min_row, min_col = target_info["origin"]
        max_row, max_col = target_info["extents"]
        col_names = target_info["col_names"]
        row_names = target_info["row_names"]

        # Get the first and last column and first and last row with a matching name.
        use_cols = [None, None]
        use_rows = [None, None]
        if override_col is not None:
            use_cols = [override_col-min_col, override_col-min_col]
        else:
            for col_num, c in enumerate(col_names):            
                if np.any([cur in col_name for cur in c]):
                    if use_cols[0] is None:
                        use_cols[0] = col_num
                    use_cols[1] = col_num
        if override_row is not None:
            use_rows = [override_row-min_row, override_row-min_row]
        else:
            for row_num, r in enumerate(row_names):
                if np.any([cur in row_name for cur in r]):
                    if use_rows[0] is None:
                        use_rows[0] = row_num
                    use_rows[1] = row_num

        # Offset column and row numbers by the origin
        use_cols = [None if c is None else c + min_col for c in use_cols]
        use_rows = [None if r is None else r + min_row for r in use_rows]
        if max_cols is not None and use_cols[0] is not None:
            assert max_cols >= 1, f"max_cols must be 1 or greater, received {max_cols} instead"
            use_cols[1] = min(use_cols[1], use_cols[0] + max_cols - 1)
        if max_rows is not None and use_rows[0] is not None:
            assert max_rows >= 1, f"max_rows must be 1 or greater, received {max_rows} instead"
            use_rows[1] = min(use_rows[1], use_rows[0] + max_rows - 1)
            
        # Create the Excel ranges
        start = get_column_letter(use_cols[0])
        if use_rows[0] is not None:
            start = "{}{}".format(start, use_rows[0])
        if use_rows[0] != use_rows[1] or use_cols[0] != use_cols[1]:
            end = get_column_letter(use_cols[1])
            if use_rows[1] is not None:
                end = "{}{}".format(end, use_rows[1])
        else:
            end = ""

        if end:
            rng = "{}:{}".format(start, end)
        else:
            rng = start

        if include_sheet_name:
            rng = "'{}'!{}".format(target_ws.title, rng)

        return excel_addr_to_fixed(rng, fixed_rows, fixed_cols)

    def get_named_cells(self, sheet_name, row_name, col_name, fixed_rows=True, fixed_cols=True, include_sheet_name=False, override_row=None, override_col=None, match_data=None):
        """Get a list of all Excel cell addresses in a worksheet that have the specified row names/columns.

        Parameters
        ----------
        sheet_name : str
            The name of the sheet to get the cells from. This is the name associated with the worksheet info.
        row_name : str | list[str]
            The row name(s) to match.
        col_name : str | list[str]
            The column name(s) to match.
        fixed_rows : bool
            If True then get the addresses with fixed rows format (eg. A$5)
        fixed_cols : bool
            If True then get the addresses with fixed columns format (eg. $A5)
        include_sheet_name : bool
            If True, then include the sheet name in the returned addresses (eg. 'Main'!A5)
        override_row : int
            If not None then for each matched cell, get the address with this as the 1-based row (instead of the row of the matched cell).
        override_col : int
            If not None then for each matched cell, get the address with this as the 1-based column (instead of the column of the matched cell).
        match_data : dict
            If set, then only find cells that has an attached_data pd.DataFrame row (from WWMeasure) that has at least one matching
            key:value pair in this dictionary.

        Returns
        -------
        list[str]
            A list of Excel addresses for all matching cells.
        """
        row_name = row_name or []
        col_name = col_name or []
        if isinstance(row_name, str):
            row_name = [row_name]
        if isinstance(col_name, str):
            col_name = [col_name]

        # Add the sheet name to all column/row names, since that is how we store the names internally
        col_name = [add_sheet_name_to_colrow_name(sheet_name, c) for c in col_name]
        row_name = [add_sheet_name_to_colrow_name(sheet_name, r) for r in row_name]

        target_ws, target_info = self.get_worksheet_and_info(sheet_name)

        min_row, min_col = target_info["origin"]
        max_row, max_col = target_info["extents"]
        col_names = target_info["col_names"]
        row_names = target_info["row_names"]

        # Get the first and last column and first and last row
        if override_col is not None:
            col_names = [[col_name]]
        if override_row is not None:
            row_names = [[row_name]]
        use_cells = []
        for col_num, c in enumerate(col_names):
            if np.any([cur in col_name for cur in c]) or override_col is not None:
                for row_num, r in enumerate(row_names):
                    if np.any([cur in row_name for cur in r]) or override_row is not None:
                        if override_col is not None:
                            col_num = override_col-min_col
                        if override_row is not None:
                            row_num = override_row-min_row
                        use_cells.append((row_num, col_num))

        use_cells = [(c[0]+min_row, c[1]+min_col) for c in use_cells]
        use_cells = ["{}{}".format(get_column_letter(c[1]), c[0]) for c in use_cells]

        # If match_data dictionary is provided, only use the cells in use_cells that has an
        # attached_data member that matches at least one key:value pair.
        ws, info = self.get_worksheet_and_info(sheet_name)
        if match_data is not None:
            def _matches_data(c):
                for key,val in match_data.items():
                    cell = ws[c]
                    attached_value = self.get_cell_attached_data_value(cell, key, None)
                    if attached_value == val:
                        return True
            
            # Only use cells where the attached_data (rows in the data) matches
            use_cells = [c for c in use_cells if _matches_data(c)]

        if include_sheet_name:
            use_cells = ["'{}'!{}".format(target_ws.title, c) for c in use_cells]
        use_cells = [excel_addr_to_fixed(c, fixed_rows, fixed_cols) for c in use_cells]

        return use_cells

    def copy_cell(self, source_cell, target_cell=None, override_values=None, target_worksheet=None):
        """Copy the specified template source_cell to the specified target_cell, translating any formulas if
        their Excel addresses are different.

        Parameters
        ----------
        source_cell : openpyxl.cell.cell.Cell
            The required source cell to copy from.
        target_cell : openpyxl.cell.cell.Cell
            The optional target cell to copy to. If None then the cell in target_worksheet with the same
            Excel address as source_cell is used.
        override_values : dict
            Any attribute values in source_cell that we want to override. If the key exists in override_values for
            an attribute name, then that value is used.
        target_worksheet : openpyxl.worksheet.worksheet.Worksheet
            If target_cell is None, then we use the cell in this worksheet at the same Excel address as source_cell as
            the target_cell.
        
        Returns
        -------
        openpyxl.cell.cell.Cell
            The target_cell that we copied to.
        """
        if target_cell is None:
            target_cell = Cell(worksheet=target_worksheet, column=source_cell.column, row=source_cell.row)

        def _source_value(key):
            # Get the key from the source_cell attributes. If the key exists in override_values then use that one instead.
            return override_values[key] if override_values is not None and key in override_values else getattr(source_cell, key, None)

        # Copy the cell value and the attached_data (ie. the DataFrame rows from WWMeasure that has been attached to the cell)
        value = _source_value("value")
        if getattr(source_cell, "attached_data", None) is not None:
            target_cell.attached_data = copy(source_cell.attached_data)
        if isinstance(value, str):
            # Remove all $ from Excel coordinates (eg. "$G3"), otherwise when we translate the formula
            # to the new location it won't properly update the coordinates in the cell value.
            value = re.sub("(?<![A-Za-z0-9])\$(?=[A-Za-z])", "", value)
            value = re.sub("(?<![0-9])\$(?=[0-9])", "", value)
            target_cell.value = Translator(value, _source_value("coordinate")).translate_formula(target_cell.coordinate)
        else:
            target_cell.value = value

        # Copy over the cell styles if there are any
        if source_cell.has_style:
            styles = [s for s in self.output_wb._named_styles if s.name == source_cell.style]
            if len(styles) > 0:
                target_cell.style = styles[0]
            target_cell.fill = copy(_source_value("fill"))
            target_cell.font = copy(_source_value("font"))
            target_cell.border = copy(_source_value("border"))
            target_cell.number_format = copy(_source_value("number_format"))
            target_cell.protection = copy(_source_value("protection"))
            target_cell.alignment = copy(_source_value("alignment"))
        return target_cell

    def get_worksheet_and_info(self, worksheet):
        """Get the OpenPYXL worksheet and our own internal worksheet info.

        Parameters
        ----------
        worksheet : str | openpyxl.worksheet.worksheet.Worksheet
            If a string, then the worksheet name that was used to add the worksheet to our info (ie. stored
            in the "sheet_name" member of the worksheet info dictionary, which might not by the same as the
            worksheet's actual title). If an OpenPYXL Worksheet object, then we retrieve the worksheet info
            associated with the Worksheet object.

        Returns
        -------
        ws : openpyxl.worksheet.worksheet.Worksheet
            The Worksheet object.
        info : dict
            The dictionary containing all the info for the worksheet, such as the origin, extents, sheet name,
            column names, row names, etc.
        """
        if isinstance(worksheet, str):
            info = [c for c in self.worksheets_info if c["sheet_name"].lower() == worksheet.lower()]
        else:
            info = [c for c in self.worksheets_info if c["ws"] == worksheet]
        if len(info) == 0:
            return None, None
        return info[0]["ws"], info[0]

    def get_all_cal_worksheets_and_info(self):
        """Get all the Worksheets and worksheet info for all calibration curves. This will exclude the main sheet.

        Returns
        -------
        list(tuple(openpyxl.worksheet.worksheet.Worksheet, dict))
            List of tuples, each tuple has element 0 being the openpyxl.worksheet.worksheet.Worksheet object and
            element 1 being the worksheet info dictionary (has the origin, extents, sheet name, column names, row names, etc.
            for the worksheet)
        """
        all = []
        for info in self.worksheets_info:
            if info["sheet_name"] == MAIN_SHEET:
                continue
            all.append((info["ws"], info))
        return all


    def parse_cell(self, cell, target_cell, data, target_sheet_name):
        """Parse a cell from the template. We'll find all tags in curly braces and replace them with values. We also
        find all custom functions (that begin with __), and either call that function immediately or add
        it as a late binder that we'll call later.

        Parameters
        ----------
        cell : openpyxl.cell.cell.Cell
            The cell as it is found in the template.
        target_cell : openpyxl.cell.cell.Cell
            The target cell that will receive the contents of cell, along with formatting and styles.
        data : pd.DataFrame
            The data associated with the cell. This is the DataFrame for the current row (eg. all data for a specific
            sample ID),
        target_sheet_name : str
            The name of the target sheet for the target_cell. This is the name in the "sheet_name" member of
            the worksheet, in our worksheets_info member.
        """
        if data is None:
            # return cell.value, None
            return None

        v = cell.value
        attached_data = []
        try:
            v, attached_data = parse_values(v, self.row_data)
        except Exception as e:
            v = str(e)            

        results = {
            "value" : v,
        }
        # Check for custom formulas
        delayed = []

        if isinstance(v, str):
            v = v.strip()
            next_pos = 0
            while True:
                ran, v, next_pos = self.run_custom_func(v, next_pos, cell, target_cell, data, target_sheet_name, None, True)
                if not ran:
                    break

        # Try to convert the value to other types
        results = {
            "value" : self.try_to_cast(cell, v)
        } 
        self.copy_cell(cell, target_cell, override_values=results)
        if len(attached_data) > 0:
            cur_attached_data = self.get_cell_attached_data(target_cell, [])
            cur_attached_data.extend(attached_data)
            self.set_cell_attached_data(target_cell, cur_attached_data)

    def run_custom_func(self, v, start_pos, cell, target_cell, data, target_sheet_name, func_name, delay_late_binders):
        """Run a custom function that is found in the specified value (v), or if it is a late binder then add it to our list
        of late binders to call later.

        Parameters
        ----------
        v : any
            The value of a cell that we are parsing. If this is a string, then we look for the func_name in this string and call
            it (or add to the list of late binder), along with the parameters specified in the string.
        start_pos : int
            The 0-based string index in v where we start to search for the func_name to execute.
        cell : openpyxl.cell.cell.Cell
            The template cell that we originally retrieved the value (v) from.
        target_cell : openpyxl.cell.cell.Cell
            The target cell that is receiving the value from the template cell.
        data : pd.DataFrame
            The data for the current row we are parsing in the output Excel file. This gets passed to the custom
            function, in case it needs to retrieve data from it.
        target_sheet_name : str
            The sheet name that target_cell belongs to. This is the "sheet_name" member of the worksheet's info object.
        func_name : str
            The name of the function, which begins with __ and has an entry in CUSTOM_FUNCS. We will search the
            string v for the next occurrence of func_name.
        delay_late_binders : bool
            If True, then if the function is a late binder (ie. "bind" in CUSTOM_FUNCS != 0) then do not call it, and
            instead add the function to the list of late binding custom functions to call later on. If False then
            we call the function immediately whether it is a late binder or not.

        Returns
        -------
        ran : bool
            If True then we either called the custom function and updated the string v with the results of the function,
            or we added the function to the list of late binders to call later on. If false then we did not
            call the custom function or add it to the late binders because there was no function in v to call.
        v : any
            The new value, which is the input parameter v with the custom function call replaced by the function's return
            value. If ran is False v is left unmodified.
        next_pos : int
            This is the next string position index in v to continue parsing from. This is the position immediately
            after the function that was called (after it is replaced with the function's return value).
            For example, if we call __FUNC from the string "test;__FUNC(a,b)after", and __FUNC returns "1", then
            the returned v is "test;1after" and next_pos is 6 (the index of the string "after").

        """
        if not isinstance(v, str):
            return False, v, start_pos
        match_func_name = func_name or "__[A-Za-z][A-Za-z0-9_]*"
        _match = re.search(CUSTOM_FUNCS_REGEX % (match_func_name), v[start_pos:])
        if _match is None:
            return False, v, start_pos
        full_match = _match[0]
        groups = list(_match.groups())
        func_name = groups[1]
        args = [strip_quotes(v).strip() for v in groups[2].split(",")]
        next_pos = _match.start() + (1 if groups[0] and len(groups[0]) > 0 else 1) + start_pos

        try:
            replace = ""
            funcinfo = CUSTOM_FUNCS[func_name] if func_name in CUSTOM_FUNCS else None 
            match_str = full_match
            if groups[0] and len(groups[0]) > 0 and groups[0][0] != CUSTOM_FUNC_SEP:
                match_str = match_str[1:]

            if funcinfo is not None and delay_late_binders and funcinfo["bind"] != 0:
                # We use target_cell for the source_cell parameter (instead of cell), since by the time the late binder is called the template will be
                # target_cell (which receives the contents of the template_cell from a call to copy_cell)
                self.add_late_binder(target_cell, target_cell, "UNK_SHEET!", funcinfo["bind"], target_sheet_name, data, func_name, self.row_data, args)
            else:
                if funcinfo is not None:
                    replace = getattr(custom_functions, funcinfo["func"])(self, *args, template_cell=cell, target_cell=target_cell, target_sheet_name=target_sheet_name)

                if isinstance(v, str):
                    v = v.replace(match_str, str(replace), 1)
        except Exception as e:
            v = str(e)
            print(f"WARNING: Exception running custom function {func_name}, will use value \"{v}\"")
            print(f"Exception for custom function warning is: {e}")
        
        return True, v, next_pos

    def try_to_cast(self, cell, value):
        """Try to cast the value in a cell to a suitable format. We will try
        to cast to a date (if the number format is for a date) or a float. If we can't
        we will simply return the original value. The actual cell is not changed, just the
        cast value is returned.
        """
        try:
            if cell.number_format and is_date_format(cell.number_format):
                value = pd.to_datetime(value)
            elif isinstance(value, str) and value.lower() in [ "true", "false"]:
                value = bool(value)
            elif value is not None:
                value = float(value)
        except:
            pass
        return value

    def handle_late_binders(self, inner=True):
        """Call all late binding custom Excel functions (specified in CUSTOM_FUNCS), in the
        order of their "bind" parameter. Late binders all have a bind value != 0, and are only called
        once the whole spreadsheet is generated for output. Early binders have bind == 0, and are called
        as soon as they are encountered when parsing the template and populating the output.
        """
        self.late_binders.sort(key=lambda x: abs(x["bind"]))
        prev_data = None
        init_row_data = self.row_data
        for late_binder in self.late_binders:
            if (inner and late_binder["bind"] < 0) or (not inner and late_binder["bind"] > 0):
                continue
            source_cell = late_binder["source_cell"]
            target_cell = late_binder["target_cell"]
            target_sheet_name = late_binder["target_sheet_name"]
            func_name = late_binder["func_name"]
            row_data = late_binder["row_data"]
            data = late_binder["data"]
            self.row_data = row_data

            ran, v, next_pos = self.run_custom_func(target_cell.value, 0, source_cell, target_cell, data, target_sheet_name, func_name, False)
            if ran:
                target_cell.value = self.try_to_cast(target_cell, v)
            
        self.row_data = init_row_data

        # self.late_binders = []
        self.late_binders = [b for b in self.late_binders if (inner and b["bind"] < 0) or (not inner and b["bind"] > 0)]

    def add_late_binder(self, source_cell, target_cell, sheet_name, bind, target_sheet_name, data, func_name, row_data, args):
        """Add a late binder, which is a custom Excel function (specified in CUSTOM_FUNCS) with bind != 1 that will be called
        only after populating the full output sheet from the template file.
        """
        self.late_binders.append({
            "bind" : bind,
            "source_cell" : source_cell,
            "target_cell" : target_cell,
            "sheet_name" : sheet_name,
            "target_sheet_name" : target_sheet_name,
            "func_name" : func_name,
            "row_data" : deepcopy(row_data),
            "args" : deepcopy(args),
            "data" : data,
        })
        
    def set_row_data(self, data):
        """Add data for parsing the current row. The values in the data can be referenced in the template spreadsheet in
        tags (ie. values in curly braces).
        
        Args:
            data (dict): Dictionary of all data. The format is:
                    "mainTarget" : {
                        "sample" : sample_df,       # Contains sample data such as ID, target, settled solids, total mass, etc.
                        "qpcr" : qpcr_df,           # Contains qpcr data such as ct_0, ct_1, ct_2
                        "..." : ...,
                    },
                    "otherTargetA" : {
                        ...
                    }
                The values (sample_df, qpcr_df) are usually pd.DataFrames, but can be scalars as well. To specify the ct_0 value
                for mainTarget, the tag in the template would be {mainTarget>qpcr>ct_0}.                
        """
        self.row_data = data.copy()

    def copy_to_position(self, source_ws, source_row, target_sheet_name, data, target_row=None, target_col=None, row_name=None):
        """Copy the specified row from the template source_ws to the end of the target sheet.

        Parameters
        ----------
        source_ws : openpyxl.worksheet.worksheet.Worksheet
            The source template worksheet we are copying the row from.
        source_row : tuple[openpyx.cell.cell.Cell]
            The row cells to copy. This is the row taken from the template source_ws, but excluding
            the first column which is meta data with row names.
        target_sheet_name : str
            The name of the target sheet to copy the row to. We copy it to the next row after the last one or target_row if specified.
            The sheet name is the one defined in our "sheet_name" member of the worksheets_info dictionaries.
        data : pd.DataFrame
            The data for the row, from WWMeasure.
        target_row : int
            The 1-based target row to copy to. If None then the next empty row is used.
        target_col : int
            The 1-based target column to copy to. If None then column 1 is used.
        row_name : str
            A row name to associate with the new row.

        Returns
        -------
        next_row : int
            The next empty row in the sheet (1-based).
        next_col : int
            The next empty column in the sheet (1-based).
        """
        target_ws, target_info = self.get_worksheet_and_info(target_sheet_name)
        if target_row is None:
            target_row = target_info["extents"][0] #target.max_row
        if target_col is None:
            target_col = 1

        rows = [source_row]

        max_col = target_col
        row_name = add_sheet_name_to_colrow_name(target_sheet_name, row_name)
        row_names = target_info["row_names"]
        row_data = target_info["row_data"]
        row_origin = target_info["origin"][0]

        # Get all columns that are merged in the template, so we can merge them in the output
        row_merged_col_ranges = []
        for cell in source_row:
            if isinstance(cell, MergedCell):
                col_num = source_row[0].column
                row_num = source_row[0].row
                for rng in source_ws.merged_cells.ranges:
                    min_row, min_col = rng.min_row, rng.min_col
                    max_row, max_col = rng.max_row, rng.max_col
                    # Only merge horizontally merged cells. We'll ignore the vertical ones
                    if min_row == max_row and row_num == min_row:
                        row_merged_col_ranges.append((target_col + min_col - 2, target_col + max_col - 2))
                break

        for cur_row in rows:
            # Add the row_name to the worksheet's row_names info
            if row_name is not None:
                cur_idx = target_row - row_origin
                while len(row_names) <= cur_idx:
                    row_names.append([])
                row_names[cur_idx].append(row_name)
            if row_data is not None:
                cur_idx = target_row - row_origin
                while len(row_data) <= cur_idx:
                    row_data.append([])
                row_data[cur_idx].append(data)
            
            # Copy all cells in the row
            for cur_col, cell in enumerate(cur_row):
                # if col == 0:
                #     continue
                cur_cell_col = target_col+cur_col
                cur_cell_row = target_row
                max_col = max(max_col, cur_cell_col)
                target_addr = "{}{}".format(get_column_letter(cur_cell_col), cur_cell_row)
                target_cell = target_ws[target_addr]
                if isinstance(target_cell, Cell):
                    results = self.parse_cell(cell, target_cell, data, target_sheet_name)

            # Merge cells in row
            if len(row_merged_col_ranges) > 0:
                for min_col, max_col in row_merged_col_ranges:
                    merge_range = "{}{}:{}{}".format(get_column_letter(min_col), target_row, get_column_letter(max_col), target_row)
                    target_ws.merge_cells(merge_range)

            target_row += 1

        return (target_row, max_col+1)
        
    def copy_widths(self, template_ws, target_sheet_name, origin_col):
        """Copy the widths of all columns in the template worksheet.

        The template worksheet has as its first row and first column all the names of the rows/columns,
        so we do not copy the widths of those.

        Parameters
        ----------
        template_ws : openpyxl.worksheet.worksheet.Worksheet
            The template worksheet to copy column widths from.
        target_sheet_name : str
            The name of the target worksheet to copy column widths to. This is the "sheet_name" member
            of our worksheets_info list of dictionaries.
        origin_col : int
            The 1-based column number which is the origin of the target worksheet. Column number 2 in
            template_ws is matched with column #origin_col in the target worksheet.
        """
        ws, target_info = self.get_worksheet_and_info(target_sheet_name)
        max_column = template_ws.max_column
        for col in range(2, max_column+1):
            dim = template_ws.column_dimensions[get_column_letter(col)]
            ws.column_dimensions[get_column_letter(col-1 + origin_col-1)].width = dim.width

    def get_columns_from_template(self, ws):
        """Get all column names from the template. These are the cell contents of the second row, starting
        at column 2, of the template (ws).
        
        For templates that output a very simple spreadsheet where the first row are column headers, this
        will correspond to the resulting column names that we would typically get when loading the output
        with Pandas. For other styles of templates, the column names retrieved here would have no use.
        """
        max_col = get_column_letter(ws.max_column)
        columns = [c.value for c in ws[f"B2:{max_col}2"][0]]
        return columns

    def copy_rows(self, template_ws, row_ids, target_sheet_name, data, target_row=None, target_col=None):
        """Copy all rows from the template worksheet, with a row id/name in row_ids, to the target output worksheet, for a single
        set of data (typically from the same sample ID).

        Parameters
        ----------
        template_ws : openpyxl.worksheet.worksheet.Worksheet
            The template worksheet to copy rows from. We will parse each row while copying.
        row_ids : str | list[str]
            The row IDs to copy. These match with the row names specified in column A of the template (eg. "main_row_data").
        target_sheet_name : str
            The target sheet name to copy to. This is the "sheet_name" member in our worksheets_info list of dictionaries. They
            are either the main sheet or one of the calibration curve sheets.
        data : pd.DataFrame
            The rows in WWMeasure associated with the current set of template rows we're copying. This is typically all the Ct
            data for the current sample ID.
        target_row : int
            The 1-based target row to copy to. If None then the next empty row is used.
        target_col : int
            The 1-based target column to copy to. If None then column 1 is used.

        Returns
        -------
        next_row : int
            The next empty row in the sheet (1-based).
        next_col : int
            The next empty column in the sheet (1-based).
        """
        data = data.sort_values(self.config.input.index_col)
        ws, target_info = self.get_worksheet_and_info(target_sheet_name)

        if isinstance(row_ids, str):
            row_ids = [row_ids]
        row_ids = [add_sheet_name_to_colrow_name(target_sheet_name, r) for r in row_ids]
        
        max_col = 1

        # Go through each row of the template, find any row that has a name in row_ids
        for row_num, row in enumerate(template_ws.iter_rows()):
            # Skip column names
            if row_num == 0:
                continue
            # Column A in the template contains the names for the row.
            cur_id = add_sheet_name_to_colrow_name(target_sheet_name, (row[0].value or "").strip().lower())
            if cur_id in row_ids:
                (last_row, last_col) = self.copy_to_position(template_ws, row[1:], target_sheet_name, data, target_row=target_row, target_col=target_col, row_name=row[0].value)
                target_row = last_row
                max_col = max(max_col, last_col)

                extents = target_info["extents"]
                target_info["extents"] = (max(target_row, extents[0]), max(max_col, extents[1]))

        return target_row, max_col

    def consolidate_extents(self):
        """Go through all of the named sheets in our output, for any of those sheets that share
        the same openpyxl.worksheet.worksheet.Worksheet object, combine their extents so that the
        extent includes all shared sheets.

        For example, the calibration sheets might actually be on the main sheet. In this case the
        main sheet extents should include all of its own data including the data of all calibration
        sheets that are on the main sheet.

        This should be called after create_main and create_calibration are called, which is when
        we know what the full extents will be.
        """
        worksheets = set([info["ws"] for info in self.worksheets_info])

        for ws in worksheets:
            matches = [info for info in self.worksheets_info if info["ws"] == ws]
            max_extents = copy(list(matches[0]["extents"]))
            for match in matches:
                max_extents[0] = max(max_extents[0], match["extents"][0])
                max_extents[1] = max(max_extents[1], match["extents"][1])
            for match in matches:
                match["extents"] = max_extents

    def get_or_create_sheet(self, sheet_name, sheet_origin, row_spacing):
        """Get a named Excel worksheet or create one if it does not exist in our output workbook.

        This will not add to our internal worksheets_info object if the sheet is newly created, that must be
        done separately.

        Parameters
        ----------
        sheet_name : str
            The title of the Excel sheet to get or create.
        sheet_origin : list[int, int]
            The sheet origin to use. See "origin" in the Returns section.
        row_spacing : int
            The spacing between rows, to add to the origin. See "origin" in the Returns section.

        Returns
        -------
        ws : openpyxl.worksheet.worksheet.Worksheet
            The Worksheet, either the existing one that was retrieved, or a new one if it didn't already exist.
        origin : list[int, int]
            The 1-based origin of the worksheet for adding new data. If the worksheet already exists then we take the extents
            of the existing sheet, advance to the next row by adding row_spacing, and use the column from sheet_origin[1].
            If a new sheet is created then the origin is simply sheet_origin. This represents where new data should be added
            to the sheet.
        extents : list[int, int]
            The 1-based extents of the part of the worksheet where we're adding new data. This will always be the same
            as origin. In later operations the extents will slowly be grown to include all newly added data.
        """
        if sheet_name in self.output_wb.sheetnames:
            ws = self.output_wb[sheet_name]
            origin = [ws.max_row + (row_spacing + 1 if ws.max_row > 1 else 0), sheet_origin[1]]
            extents = origin
        else:
            ws = self.output_wb.create_sheet(sheet_name)
            origin = sheet_origin
            extents = origin

        return ws, origin, extents

    def get_all_paired_cal_sheets(self, target_a, target_b, no_qaqc_only=True):
        """Get all calibration curve sheets where one curve is for target_a and the other for
        target_b.

        Parameters
        ----------
        target_a : str
            The first target to retrieve.
        target_b : str
            The second target to retrieve, that we'll pair each curve from target_a with.
        no_qaqc_only : bool
            If True, then only retrieve curves where QAQC has not yet been performed. These
            are the sheets where the worksheets_info object has "ran_qaqc" set to False.

        Returns
        -------
        list[tuple(dict, dict)]
            List of all paired sheets. The first item in each tuple is the sheet with target_a, the second the sheet with target_b.
            These sheets will typically be on the same plate, and can be used to compare the two calibration curves.
        """
        pairs = []

        # Get all sheets for target_a and target_b.
        target_a_sheets = self.get_cal_sheets_by_target(target_a, no_qaqc_only=no_qaqc_only)
        target_b_sheets = self.get_cal_sheets_by_target(target_b, no_qaqc_only=no_qaqc_only)
        if len(target_a_sheets) == 0 or len(target_b_sheets) == 0:
            return pairs

        # Go through all target_a sheets. Find a paired target_b sheet on the same plate if possible.
        # If there is no target_b curve on the same plate, then pair by matching the index into
        # target_a_sheets and target_b_sheets
        for idx_a, sheet_a in enumerate(target_a_sheets):
            match_sheet_b = None
            # Go through each of target_b_sheets, find a matching plateID
            for sheet_b in target_b_sheets:
                if sheet_b["plateID"] == sheet_a["plateID"]:
                    match_sheet_b = sheet_b
                    break
            if match_sheet_b == None:
                # No target_b curve with the same plateID, so instead pair by index into target_a_sheets and
                # target_b_sheets
                match_sheet_b = target_b_sheets[min(idx_a, len(target_b_sheets)-1)]
            if match_sheet_b is not None:
                pairs.append((sheet_a, match_sheet_b))
        return pairs

    def get_cal_sheets_by_target(self, target, no_qaqc_only=True):
        """Get all calibration curve sheets with the specified target.

        Parameters
        ----------
        target : str
            The target to get all calibration curves for.
        no_qaqc_only : bool
            If True, then only retrieve curves where QAQC has not yet been performed. These
            are the sheets where the worksheets_info object has "ran_qaqc" set to False.

        Returns
        -------
        list[dict]
            List of all calibration sheets info, for all sheets for the specified target.
        """
        target = target.lower()
        sheets = []
        for info in self.worksheets_info:
            if info["sheet_name"] != MAIN_SHEET and info["target"].lower() == target:
                if (no_qaqc_only and not info["ran_qaqc"]) or not no_qaqc_only:
                    sheets.append(info)
        return sheets

    def create_calibration(self, data):
        """Create all calibration curves found in the data.

        We will create all the calibration curve sheets, Excel graphs, and the calculated slopes, intercepts, R^2, efficiency,
        etc. for each calibration curve.
        """
        if self.config.template.calibration_sheet_name not in self.template_xl:
            print("WARNING: No calibration template exists in the template file. Skipping calibration")
            return
            
        # Our template sheet
        template_cal = self.template_xl[self.config.template.calibration_sheet_name]

        col_names, row_names = get_template_colrow_names(template_cal)

        # Retrieve all Std measurements, split into groups by target type (eg. N1, N2, Pepper)
        std_data = data[data[self.config.input.measure_type_col] == self.config.input.measure_type_std]
        std_data = std_data.sort_values(self.config.input.index_col)
        sq_data = std_data #data[data[self.config.input.measure_type_col] == "SQ"]
        # We will iterate over each data row in the order of the unique sample IDs in
        # sq_data. Within each sample ID, we will go from largest SQ value (copies/well) to
        # smallest.
        sq_data = sq_data.sort_values(self.config.input.sq_col, ascending=False)
        target_names = [n for n in std_data[self.config.input.target_col].unique() if n]
        target_names.sort()

        calibration_location = self.config.template.calibration_location
        cal_origin = self.config.template.cal_origin
        if calibration_location == "main_sheet":
            # Put all calibration info and plots on the main sheet
            ws, ws_info = self.get_worksheet_and_info(MAIN_SHEET)
            first_row = ws_info["extents"][0] + cal_origin[0] - 1
            target_col = cal_origin[1]
        elif calibration_location == "cal_sheet" or calibration_location == "hide":
            ws, origin, extents = self.get_or_create_sheet(SINGLE_CAL_SHEET, self.config.template.cal_origin, self.config.template.rows_between_cal_groups)
            # Note: will readd cal_origin later on
            first_row = origin[0]
            target_col = origin[1]
        elif calibration_location == "caltarget_sheet":
            # Use multiple extra sheets for each separate calibration (1 calibration/sheet)
            first_row = cal_origin[0]
            target_col = cal_origin[1]

        # Origin row, col of the calibration info
        # first_row += cal_origin[0] - 1
        # target_col += cal_origin[1] - 1

        # Go through all targets that have calibration curve data.
        def _form_groups(df):
            # This is similar to doing df.groupby(self.config.input.target_col), but we use our own ordering of the groups
            # with main_targets first, then other_targets, then normalizing_targets, then inhibition_targets,
            # then any target that isn't in these groups.
            group_names = sorted(df[self.config.input.target_col].unique())
            all_groups = []
            for cur_targets in [self.config.input.main_targets, self.config.input.other_targets, self.config.input.normalizing_targets, self.config.input.inhibition_targets]:
                if cur_targets:
                    for target in cur_targets:
                        if target in group_names:
                            all_groups.append([target, None])
                            group_names.remove(target)
            
            all_groups.extend([[g, None] for g in group_names])
            for idx in range(len(all_groups)):
                all_groups[idx][1] = df[df[self.config.input.target_col] == all_groups[idx][0]]
            return all_groups
                        
        for idx, (target_name, target_df) in enumerate(_form_groups(std_data)):
            for plate_id, plate_df in target_df.groupby(self.config.input.plate_id_col):
                row_data_kwargs = {
                    "plateID" : plate_id,
                }
                
                # cur_std_data = std_data[std_data[self.config.input.target_type_col].str.lower() == target_name.lower()]
                # cur_sq_data = sq_data[sq_data[self.config.input.target_type_col].str.lower() == target_name.lower()]
                cur_std_data = plate_df
                cur_sq_data = sq_data[(sq_data[self.config.input.target_col].str.lower() == target_name.lower()) & (sq_data[self.config.input.plate_id_col] == plate_id)]

                if len(cur_std_data.index) == 0:
                    continue

                ws_title = self.get_standard_curve_id(target=target_name, plateID=plate_id)

                if calibration_location == "caltarget_sheet":
                    ws, origin, extents = self.get_or_create_sheet(ws_title, self.config.template.cal_origin, self.config.template.rows_between_cal_groups)
                    first_row = origin[0]
                    target_col = origin[1]

                target_row = first_row

                if self.get_worksheet_and_info(ws_title)[0] is None:
                    # Add the calibration sheet info since we don't have one yet
                    cur_col_names, cur_row_names = add_sheet_name_to_colrow_names(ws_title, col_names, row_names)
                    self.worksheets_info.append({
                        "sheet_name" : ws_title,
                        "ws" : ws,
                        "origin" : [target_row, target_col],
                        "extents" : [target_row, target_col],
                        "col_names" : cur_col_names,
                        "row_names" : cur_row_names,
                        "row_data" : [],
                        "target" : target_name,
                        "plateID" : plate_id,
                        "ran_qaqc" : False,
                    })
                
                _, cal_info = self.get_worksheet_and_info(ws_title)

                # Prepare row data for banner and header
                self.prepare_row_data(None, None, target_name, is_calibration_curve=True, **row_data_kwargs)

                # Copy the section banner
                target_row, _ = self.copy_rows(template_cal, "cal_row_banner", ws_title, cur_std_data, target_row=target_row, target_col=target_col)

                # Copy the header
                target_row, _ = self.copy_rows(template_cal, "cal_row_header", ws_title, cur_std_data, target_row=target_row, target_col=target_col)

                # For each sample ID of the sq_data, copy "cal_row_data". This typically corresponds to one data row in the output
                # (but depends on the template)
                idx = 0
                cal_sq = []
                cal_logsq = []
                cal_ct = []
                for sample_id in cur_sq_data[self.config.input.sample_id_col].str.lower().unique():
                    cur_data = cur_std_data[cur_std_data[self.config.input.sample_id_col].str.lower() == sample_id]
                    if len(cur_data.index) == 0:
                        continue
                    self.prepare_row_data(data, sample_id, target_name, item_number=idx, is_calibration_curve=True, **row_data_kwargs)

                    # Calculate SQ (Copies), log10(SQ), and StdAvg(Ct), so we can calculate the calibration curve
                    # parameters in memory (separate from the Excel File)
                    sq = self.row_data["mainTarget"]["sample"][self.config.input.sq_col].iloc[0]

                    ct = self.row_data["mainTarget"]["qpcr"]["ct"][:self.config.input.slope_and_intercept_replicates]# cur_data[self.config.input.value_col][:self.config.input.slope_and_intercept_replicates]#.mean()
                    ct = [c for c in ct.values if isinstance(c, (float, int))]
                    logsq = math.log10(sq)
                    cal_sq.extend([sq] * len(ct))
                    cal_logsq.extend([logsq] * len(ct))
                    cal_ct.extend(ct)

                    target_row, _ = self.copy_rows(template_cal, CAL_ROW_DATA, ws_title, cur_data, target_row=target_row, target_col=target_col)
                    idx += 1

                # If calibration_multi is set, then we prefer to use at least calibration.multi.min_points
                # values for the calibration curve. If we have more points, we pick whichever number of
                # points that result in a slope closest to calibration_multi.preferred_slope
                max_points = len(cal_logsq)
                min_points = min(self.config.input.calibration_multi.min_points, max_points)
                calibration_multi = self.config.input.get("calibration_multi", None)
                if calibration_multi is not None:
                    preferred_slope = calibration_multi.preferred_slope
                    items = [(cal_logsq[:min_points], cal_ct[:min_points])]
                    if max_points > min_points:
                        for i in range(min_points+1, max_points+1):
                            items.append((cal_logsq[:i], cal_ct[:i]))
                else:
                    preferred_slope = None
                    items = [(cal_logsq, cal_ct)]
                # Calculate all the slopes, and choose the best one
                slope = intercept = rsq = None
                num_points = max_points
                
                # print("SAMPLEID:", sample_id, len(cur_sq_data.index), items)
                
                for idx, (logsq, ct) in enumerate(items):
                    A = np.vstack([logsq, np.ones(len(logsq))]).T
                    _slope, _intercept = np.linalg.lstsq(A, ct, rcond=None)[0]
                    corr_matrix = np.corrcoef(logsq, ct)
                    _rsq = corr_matrix[0,1]**2

                    if slope is None or abs(_slope - preferred_slope) < abs(slope - preferred_slope):
                        num_points = idx + min_points
                        slope, intercept = _slope, _intercept
                        rsq = _rsq 

                eff = 10**(-1/slope) - 1

                # Add our calculated values.
                cal_info["cal_curve"] = {
                    # "reg" : reg,
                    "sq" : sq,
                    "slope" : slope,
                    "intercept" : intercept,
                    "rsq" : rsq,
                    "eff" : eff,
                    "num_points" : num_points,
                    "max_points" : max_points,
                    "max_ct" : cal_ct[-1],
                    "min_ct" : cal_ct[0],
                }
                for idx, ct in enumerate(cal_ct):
                    cal_info["cal_curve"][f"avg_std_{idx}"] = ct

                # Copy footer
                self.prepare_row_data(None, None, target_name, is_calibration_curve=True, **row_data_kwargs)
                target_row, _ = self.copy_rows(template_cal, "cal_row_footer", ws_title, cur_std_data, target_row=target_row, target_col=target_col)

                target_col = self.get_worksheet_and_info(ws_title)[1]["extents"][1]

                # Add the calibration curve chart
                if self.config.template.calibration_location != "hide":
                    chart = ScatterChart()
                    analysis_date = data[self.config.input.analysis_date_col].dropna().unique()
                    analysis_date = analysis_date[0] if len(analysis_date) > 0 else "No Date"
                    # analysis_date = cur_std_data[self.config.input.analysis_date_col].iloc[0]

                    chart.title = f"{target_name} Cal ({analysis_date})"
                    chart.legend = None
                    # rng_logct = self.get_named_range(ws_title, CAL_ROW_DATA, "cal_col_logct", include_sheet_name=True, max_rows=num_points, max_cols=num_points)
                    # rng_ct = self.get_named_range(ws_title, CAL_ROW_DATA, "cal_col_graphct", include_sheet_name=True, max_rows=num_points, max_cols=num_points)
                    rng_logct = self.get_named_range(ws_title, CAL_ROW_DATA, "cal_col_logct", include_sheet_name=True)
                    rng_ct = self.get_named_range(ws_title, CAL_ROW_DATA, "cal_col_graphct", include_sheet_name=True)

                    try:
                        # ref_logct = Reference(ws, range_string=rng_logct)
                        # ref_ct = Reference(ws, range_string=rng_ct)
                        series = Series(rng_ct, rng_logct,)
                        series.graphicalProperties.line.noFill = True
                        series.marker.symbol = "circle"
                        series.trendline = Trendline(dispEq=True, dispRSqr=True, trendlineType="linear")
                        series.trendline.graphicalProperties = GraphicalProperties()
                        series.trendline.graphicalProperties.line.dashStyle = "sysDot"
                        series.trendline.graphicalProperties.line.solidFill = "4472C4"
                        series.trendline.graphicalProperties.line.width = 18750
                        series.trendline.trendlineLbl = TrendlineLabel(layout=Layout(manualLayout=ManualLayout(x=1, y=-1)))
                        chart.series.append(series)

                        # Set chart width and height
                        # h = points_to_cm(self.calc_rows_height(ws, first_row, target_row-1))
                        h = points_to_cm(self.calc_rows_height(ws, first_row, min(first_row + self.config.template.max_chart_height_rows - 1, target_row-1)))
                        chart.height = h
                        chart.width = h*2.25
                        ws.add_chart(chart, "{}{}".format(get_column_letter(target_col+self.config.template.chart_column_spacing), first_row))

                        # Get a rough estimate of the column of the right edge (plus some padding)
                        # of the chart, so we know where the next column is. We can only get a guess,
                        # since the width of columns in Excel is defined in characters, which depends on the
                        # font being used.
                        w = 0
                        target_col = target_col+self.config.template.chart_column_spacing
                        chart_width_chars = estimated_cm_to_chars(chart.width)
                        while w < chart_width_chars + units.DEFAULT_COLUMN_WIDTH / 5:
                            cur_w = ws.column_dimensions[get_column_letter(target_col)].width
                            w += cur_w
                            target_col += 1
                        
                        extents = self.get_worksheet_and_info(ws_title)[1]["extents"]
                        self.get_worksheet_and_info(ws_title)[1]["extents"] = [extents[0], target_col]
                    except Exception as e:
                        print(f"ERROR: Could not add calibration graph: {e}")
            
    def calc_rows_height(self, ws, first_row, last_row):
        """Calculate the height of rows in an Excel spreadsheet.
        """
        h = 0
        for r in range(first_row, last_row+1):
            cur_h = ws.row_dimensions[r].height
            h += (cur_h if cur_h is not None else units.DEFAULT_ROW_HEIGHT) + 1
        return h
    
    def prepare_row_data(self, data, sample_id, main_target, *, item_number=None, is_calibration_curve=False, **kwargs):
        self.row_data = {}
        self.row_data["mainTarget"] = self.get_row_data_for_target(data, sample_id, main_target, is_calibration_curve=is_calibration_curve)
        self.row_data["itemNumber"] = item_number
        if not is_calibration_curve:
            all_targets_info = [
                {
                    "tag" : "otherTarget{letter}",
                    "targets" : self.config.input.other_targets,
                },
                {
                    "tag" : "normTarget{letter}",
                    "targets" : self.config.input.normalizing_targets,
                },
                {
                    "tag" : "inhTarget{letter}",
                    "targets" : self.config.input.inhibition_targets,
                },
            ]
            for info in all_targets_info:
                tag = info["tag"]
                targets = info["targets"] or []
                for idx, cur_target in enumerate(targets):
                    letter = get_column_letter(idx+1).upper()
                    cur_tag = tag.format(letter=letter)
                    self.row_data[cur_tag] = self.get_row_data_for_target(data, sample_id, cur_target)
        for key, val in kwargs.items():
            self.row_data[key.lower()] = val
        self.row_data["mainTargetName"] = main_target
        return self.row_data
    
    def get_row_data_for_target(self, data, sample_id, target_name, is_calibration_curve=False):
        # if not sample_id:
        #     return {}

        if sample_id and data is not None:
            sample_data = data[data[self.config.input.sample_id_col].str.lower() == sample_id.lower()]
        else:
            sample_data = data

        sample_df = pd.DataFrame(columns=["analysisDate", "sampleDate", "sampleID", "siteID", "typeShortDescription", "typeDescription", "totalVolume", "emptyTubeMass", "totalTubeMass", "settledSolids", "extractedMass", "sq", "plateID", "standardCurveID"])
        qpcr_df = pd.DataFrame(columns=["ct", "ct_avg"])
        methods_df = pd.DataFrame(columns=self.methods.methods_df.columns)

        if sample_data is not None:
            if target_name is not None:
                target_sample_df = sample_data[sample_data[self.config.input.target_col].str.lower() == target_name.lower()]
                sample_df = target_sample_df[sample_df.columns].drop_duplicates(self.config.input.sample_id_col)
                methods_df = self.methods.get_row_for_target(target_name)

                # Create qpcr_df
                # Ct
                target_filt = sample_data[self.config.input.target_col].str.lower() == target_name.lower()
                if is_calibration_curve:
                    std_filt = sample_data[self.config.input.measure_type_col] == self.config.input.measure_type_std
                    ct_data = sample_data[std_filt & target_filt] # select_data(sample_data, self.config.input.unit_other_col, "Ct_Std", match_target=target_name, drop_duplicates_by_sample_id=False)
                else:
                    unk_filt = sample_data[self.config.input.measure_type_col] == self.config.input.measure_type_unknown
                    ct_data = sample_data[unk_filt & target_filt] # select_data(sample_data, self.config.input.unit_col, "Ct", match_target=target_name, drop_duplicates_by_sample_id=False)
                values = ct_data[self.config.input.ct_col]
                outliers = ct_data[OUTLIER_COL]

                if isinstance(values, pd.Series):
                    values = values.tolist()
                values = list(values)
                if isinstance(outliers, pd.Series):
                    outliers = list(outliers)
                outliers = list(outliers) if outliers is not None else None

                # df = pd.DataFrame(columns=["val", "group"])
                df_values = pd.DataFrame(columns=["val"])
                for i in range(len(values)):
                    # Add the value, or instead add the outlier if it exists
                    val = "" if i >= len(values) else (values[i] if outliers is None or pd.isna(outliers[i]) else f'[{outliers[i]}]')
                    # cur_data = None if data is None or i >= len(data.index) else data.iloc[i]

                    # Try to convert to float, use None if it can't be converted
                    try:
                        float_val = None
                        float_val = float(val)
                    except:
                        pass
                    df_values = df_values.append({"val" : float_val}, ignore_index=True)
                    qpcr_df.loc[i, "ct"] = val

                # Add the average
                if not qpcr_df.empty:
                    qpcr_df["ct_avg"] = df_values["val"].mean()
                   
        target_name_nodil, dilution = self.methods.split_target_and_dilution_factor(target_name)        
        dilution_text = f"1/{dilution}" if dilution != 1 else "Full"
        dilution_text_short = f"{dilution}" if dilution != 1 else "Full"
        target_name_withdilution = target_name_nodil if dilution == 1 else f"{target_name_nodil}:{dilution}"

        row_data = {
            "sample" : sample_df,
            "qpcr" : qpcr_df,
            "method" : methods_df,
            "targetName" : target_name,
            "targetNameNoDilution" : target_name_nodil,
            "targetNameWithDilution" : target_name_withdilution,
            "dilutionFactor" : dilution,
            "dilutionText" : dilution_text,
            "dilutionTextShort" : dilution_text_short,
        }

        return row_data.copy()

    def get_recognized_target(self, target_name, main_targets_only=False):
        """Map the specified target name to a unique one, ensuring that target names are consistent.
        The map is specified by input.main_targets and input.other_targets in the config file.
        """
        if target_name is None:
            return None
        recognized_targets = self.config.input.get("main_targets", [])
        if isinstance(recognized_targets, str):
            recognized_targets = [recognized_targets]

        # Add other targets
        if not main_targets_only:
            other_targets = self.config.input.get("other_targets", []) or []
            if isinstance(other_targets, str):
                other_targets = [other_targets]
            inhibition_targets = self.config.input.get("inhibition_targets", []) or []
            if isinstance(inhibition_targets, str):
                inhibition_targets = [inhibition_targets]
            normalizing_targets = self.config.input.get("normalizing_targets", []) or []
            if isinstance(normalizing_targets, str):
                normalizing_targets = [normalizing_targets]
            recognized_targets = list(recognized_targets) + list(other_targets) + list(inhibition_targets) + list(normalizing_targets)
            
        if len(recognized_targets) == 0:
            return target_name
        
        lower_recognized_targets = [g.lower() for g in recognized_targets]

        if target_name.lower() in lower_recognized_targets:
            return recognized_targets[lower_recognized_targets.index(target_name.lower())]
        
        return None
    
    def create_main(self, group_name, data):
        """Create the main sheet for the specified group of data. The data should usually represent a full QPCR run, possibly
        from multiple plates running the same samples.

        Returns
        -------
        bool
            True if the main sheet has data, False if it doesn't.
        """
        ct_values = data[data[self.config.input.measure_type_col] == self.config.input.measure_type_unknown]
        # _target_groups = ct_values.groupby(self.config.input.target_type_col)
        target_groups = []

        # Group by all the main_targets (specified in config file)
        other_targets = [g.lower() for g in self.config.input.other_targets or []]
        for target_name in self.config.input.main_targets:
            target_filt = ct_values[self.config.input.target_col].str.lower() == target_name.lower()
            if target_filt.sum() == 0:
                continue
            # Collect all other targets that we're interested in (other_targets in the config file)
            other_targets_filt = ct_values[self.config.input.target_col].str.lower().isin(other_targets)
            group = ct_values[target_filt | other_targets_filt]
            target_groups.append((target_name, group))

        if len(target_groups) == 0:
            return False

        main_ws, info = self.get_worksheet_and_info(MAIN_SHEET)
        freeze_panes = self.config.template.get("main_sheet_freeze_panes", None)
        if freeze_panes:
            main_ws.freeze_panes = freeze_panes

        # origin = info["origin"]
        # extents = info["extents"]
        target_row = info["extents"][0]
        target_col = info["origin"][1]

        main_banners_and_headers_once = self.config.template.get("main_banners_and_headers_once", False)
        requires_banners_and_headers = (main_banners_and_headers_once and main_ws.max_row <= 1 and main_ws.max_column <= 1) or not main_banners_and_headers_once

        template_main = self.template_xl[self.config.template.main_sheet_name]

        # Copy all column dimensions from the template to the main sheet
        self.copy_widths(template_main, MAIN_SHEET, target_col)

        col_names, row_names = get_template_colrow_names(template_main, sheet_name=info["sheet_name"])
        info["col_names"] = col_names
        info["row_names"] = row_names
        info["row_data"] = []

        row_data_kwargs = {
            "groupName" : group_name,
        }

        # Copy the banner once
        if requires_banners_and_headers:
            self.prepare_row_data(None, None, target_name, **row_data_kwargs)
            target_row, _ = self.copy_rows(template_main, "main_row_banner", MAIN_SHEET, ct_values, target_row=target_row, target_col=target_col)
        self.main_columns = self.get_columns_from_template(template_main)

        for target_name, target_group in target_groups:
            target_group = target_group[target_group[self.config.input.target_col] == target_name]

            # Copy the header
            if requires_banners_and_headers:
                self.prepare_row_data(None, None, target_name, **row_data_kwargs)
                target_row, _ = self.copy_rows(template_main, "main_row_header", MAIN_SHEET, ct_values, target_row=target_row, target_col=target_col)

            # For each sample ID of the data, copy "main_row_data"
            for idx, sample_id in enumerate(target_group[self.config.input.sample_id_col].str.lower().unique()):
                self.prepare_row_data(data, sample_id, target_name, item_number=idx, **row_data_kwargs)
                
                match_filt = target_group[self.config.input.sample_id_col].str.lower() == sample_id
                cur_data = target_group[match_filt]
                target_row, _ = self.copy_rows(template_main, MAIN_ROW_DATA, MAIN_SHEET, cur_data, target_row=target_row, target_col=target_col)

            if main_banners_and_headers_once:
                requires_banners_and_headers = False

        return True

    def remove_non_main_sheets_info(self):
        """Remove the worksheets_info objects (that contains info about all sheets, such as the Worksheet object,
        title, origin, extents, row data, etc) for all sheets that aren't the main sheet (ie. remove the Calibration sheets).
        """
        remove = []
        for info in self.worksheets_info:
            if info["sheet_name"] != MAIN_SHEET:
                # self.worksheets_info.remove(info)     
                remove.append(info)
                if self.config.template.calibration_location == "hide":
                    title = info["ws"].title
                    if title in self.output_wb.sheetnames:
                        self.output_wb.remove(info["ws"])
        for remove_info in remove:
            self.worksheets_info.remove(remove_info)

    def get_named_cell_address_or_value(self, sheet_name, id, fixed_row=False, fixed_col=False, prefer_precalculated=True, target_sheet_name=None):
        """Retrieve a named cell address, or get the value stored in the address if possible.

        The named cells are either set with the custom function __SETCELL (specified in the template). We also
        create named cells in code when calculating the standard curves. This way we can retrieve those values (eg. slope,
        intercept, R^2, ...) either as Excel addresses or as literal values, which we would want if the calibration
        curves are not included in the final output.

        Parameters
        ----------
        sheet_name : str
            The sheet name to get the cell from.
        id : str
            The cell ID in sheet_name to get.
        fixed_row : bool
            If True, then get the cell address with a fixed row (eg. A$5).
        fixed_col : bool
            If True, then get the cell address with a fixed column (eg. $A5).
        prefer_precalculated : bool
            If True, and we have the literal value available for the cell, then retrieve the precalculated value instead.
            This is for all calibration curve values. If the precalculated value does not exist then the
            cell address will be retrieved instead.
        target_sheet_name : str
            This is the sheet that will be referencing the retrieved cell address. If the worksheet
            is not the same as sheet_name, then we will include sheet_name to the returned cell address
            (eg. 'Cal'!A5).

        Returns
        -------
        str | float
            Either the cell address or the literal value if the precalculated value is available and prefer_precalculated is True.
        """
        if self.config.template.calibration_location == "hide":
            prefer_precalculated = True
        ws, info = self.get_worksheet_and_info(sheet_name)
        ws_target, _ = self.get_worksheet_and_info(target_sheet_name)
        if info is not None:
            if prefer_precalculated and id in info["cal_curve"]:
                return info["cal_curve"][id]
                
            addr = info[id]
            addr = excel_addr_to_fixed(addr, fixed_col, fixed_row)
            if ws_target != ws:
                return f"'{ws.title}'!{addr}"
            else:
                return addr

        return None

    ##########################################################################################

    def get_cell_attached_data(self, cell, default=[]):
        """Get the attached_data member of the cell. This is all pd.DataFrame rows added to the
        cell, due to a tag (eg. {value_covn1_0}) being parsed for the cell.
        """
        return getattr(cell, "attached_data", default)

    def set_cell_attached_data(self, cell, attached_data):
        """Set the attached_data member of the cell. This is all pd.DataFrame rows added to the
        cell, due to a tag (eg. {value_covn1_0}) being parsed for the cell.
        """
        setattr(cell, "attached_data", attached_data)

    def get_cell_attached_data_value(self, cell, key, default=None):
        """Get a value in the attached_data of the cell, by looking for the first attached_data that
        has the column key.
        """
        attached_data = self.get_cell_attached_data(cell, [])
        key = "{%s}" % key
        for data in attached_data:
            v, _ = parse_values(key, data)
            if v != key:
                return v
        return default

    def get_cell_plate_id(self, cell, default=None):
        """Get the plateID of the attached_data for the cell, by looking for the first attached_data that
        has the column self.config.input.plate_id_col. (this calls get_cell_attached_data_value)
        """
        # @TODO: Do not hardcode "sample>plateID"
        return self.get_cell_attached_data_value(cell, "sample>plateID", None)

    def get_cell_standard_curve_id(self, cell, default=None):
        """Get the standardCurveID of the attached_data for the cell, by looking for the first attached_data that
        has the column self.config.input.standard_curve_id_col. (this calls get_cell_attached_data_value)
        """
        # @TODO: Do not hardcode "sample>standardCurveID"!
        return self.get_cell_attached_data_value(cell, "sample>standardCurveID", None)

    def get_calibration_value(self, sheet_name, cal_value):
        """Get a calibration curve value from the specified calibration curve sheet ID, whose name
        is in the format specified by CAL_SHEET_FMT ("Cal-{targetName}-{plateID}").

        Parameters
        ----------
        sheet_name : str
            The sheet name of the calibration curve, The name is in the format specified by CAL_SHEET_FMT ("Cal-{targetName}-{plateID}").
        cal_value : str
            The calibration value to retrieve. Can be any of:
                "sq" : All starting quantity (copies/well) values for the calibration curve, from highest to lowest.
                "slope" : Slope of the curve
                "intercept" : Intercept of the curve
                "rsq" : R^2 value of the curve
                "eff" : Efficiency (decimal) of the curve
                "num_points" : Number of points used in calculating the curve. This is the same as max_points if
                    calibration_multi is not defined in the config file. If calibration_multi is specified, then the
                    number of points used is the number of points needed to get the curve's slope closest to
                    calibration_multi.preferred_slope. At least calibration_multi.min_points is used (unless fewer than
                    min points is available, in which case all available points are used)
                "max_points" : Total number of points available in calculating the curve (see num_points)
                "max_ct" : The average Ct value for the smallest copies/well sample (ie. should have the largest Ct value).
                "min_ct" : The average Ct value for the largest copies/well sample.

        Returns
        -------
        float | int
            The requested value, or None if not available.
        """
        ws, info = self.get_worksheet_and_info(sheet_name)
        if info is None or "cal_curve" not in info:
            return None

        return info["cal_curve"][cal_value]

    def make_file_splits(self, df):
        """Split the DataFrame into groups, where each group will be analyzed separately and output to a different file.
        The split depends on the output file name that is formed from target_file passed to the constructor. target_file
        optionally has tags that depend on the site ID of each sample ({site_id}, {site_title}, {parent_site_id}, {parent_site_title},
        and {sample_type}). Sites with the same output file (including the path) will be grouped together in the same file.

        Parameters
        ----------
        df : pd.DataFrame
            The dataframe to split.
        
        Returns
        -------
        list[tuple(dict, pd.DataFrame)]
            All groups to split up. This is a list of tuples. The first element of each tuple is a dictionary that provides
            info about the split, in the format {"parentSiteID" : str, "parentSiteTitle" : str}. The second
            element of the tuple is the DataFrame of the split. Note that there can be overlap between each group,
            since all required standard curves and additional data are included in each group.
        """
        def _clean_site_id(site_id):
            # site_id = str(site_id).replace(f"{self.config.input.lab_id}{self.config.input.lab_id_separator}", "")
            site_id = re.sub("[^A-Za-z0-9_\.-]", "", site_id)
            return site_id

        unknowns_filt = df[self.config.input.measure_type_col] == self.config.input.measure_type_unknown
        cleaned_site_id_col = "___cleaned_site_id___"
        df[cleaned_site_id_col] = df[self.config.input.site_id_col].map(_clean_site_id)
        groups = self.sites.group_by_file_template(df, cleaned_site_id_col, self.target_file, intersection_filter=unknowns_filt, always_include_filter=~unknowns_filt)
        del df[cleaned_site_id_col]
        return groups

    def make_inner_splits(self, df):
        """For each file group (from make_file_splits), split the DataFrame into additional groups by analysis date.

        Parameters
        ----------
        df : pd.DataFrame
            The DataFrame to split.
        
        Returns
        -------
        list[tuple(str, pd.DataFrame)]
            List of groups, each item is a tuple, with element 0 being the group name (ie. analysis date)
            and element 1 being the data frame for the group.
        """
        # Group by analysis date, and include all Ct_Std and SQ values needed for creating the calibration curves for the analysis date
        analysis_dates = df[self.config.input.analysis_date_col].unique()
        # We remove pd.NaT from analysis_dates before we do analysis_dates.sort() because
        # comparing NaT to datetime is deprecated
        has_nat = np.any([pd.isnull(d) for d in analysis_dates])
        analysis_dates = [d for d in analysis_dates if not pd.isnull(d)]
        analysis_dates.sort()
        if has_nat:
            analysis_dates.append(pd.NaT)
        groups = []
        for analysis_date in analysis_dates:
            # Include all items for the current analysis_date
            if pd.isna(analysis_date):
                cur_filt = df[self.config.input.analysis_date_col].isna()
            else:
                cur_filt = df[self.config.input.analysis_date_col] == analysis_date
            # Include all needed standard curve data (measureType "std")
            standard_curve_ids = df[cur_filt][self.config.input.standard_curve_id_col].unique()
            std_filt = df[self.config.input.standard_curve_id_col].isin(standard_curve_ids) & df[self.config.input.measure_type_col].isin([self.config.input.measure_type_std])
            groups.append((analysis_date, df[cur_filt | std_filt]))

        return groups

    def get_id_without_rerun(self, sample_id, target=None):
        collapse_reruns = self.config.input.get("collapse_reruns", None)
        if collapse_reruns is None:
            return sample_id
        if collapse_reruns.targets is not None and len(collapse_reruns.targets) > 0 and target is not None and target.lower() not in [g.lower() for g in collapse_reruns.targets]:
            return sample_id

        match = re.search(collapse_reruns.sample_rerun_id, sample_id)
        return match[1] if match and match[1] else sample_id

    def get_rerun_number(self, sample_id, target=None):
        collapse_reruns = self.config.input.get("collapse_reruns", None)
        if collapse_reruns is None:
            return None
        if collapse_reruns.targets is not None and len(collapse_reruns.targets) > 0 and target is not None and target.lower() not in [g.lower() for g in collapse_reruns.targets]:
            return None

        match = re.search(collapse_reruns.sample_rerun_number, sample_id)
        return int(match[1] if match[1] else 0) if match is not None else None

    def prepare_reruns(self, df):
        # @TODO: Test this and rewrite (with comments)
        if "collapse_reruns" not in self.config.input:
            return df

        reruns = pd.DataFrame()
        reruns["original_id"] = df[self.config.input.sample_id_col]
        reruns["id"] = df[self.config.input.sample_id_col].map(self.get_id_without_rerun)
        reruns["number"] = df[self.config.input.sample_id_col].map(self.get_rerun_number)
        reruns["target"] = df[self.config.input.target_col]
        reruns["measure_type"] = df[self.config.input.measure_type_col]
        reruns["index"] = df[self.config.input.index_col]
        # reruns["index"] = df[self.config.input.index_col]
        reruns.index = df.index
        reruns = reruns.dropna(subset=["number"], how="any")
        reruns["number"] = reruns["number"].astype(int)

        reruns = reruns[reruns["target"].isin(self.config.input.collapse_reruns.targets)]
        # print(reruns)

        groups = reruns.groupby(["id", "target", "measure_type"], as_index=False)
        for (group_id, group_target, group_measure_type), group_idx in groups.groups.items():
            group = reruns.loc[group_idx]
            group_id, group_target, group_measure_type = group[["id", "target", "measure_type"]].iloc[0].tolist()
            max_rerun = group["number"].max()

            # Go through all reruns, from 0 to max_rerun, project down to non-rerun values at each step
            for r in range(max_rerun+1):
                cur_reruns = group[group["number"] == r]
                if len(cur_reruns.index) == 0:
                    continue                
                # Delete the first len(cur_reruns.index) non-reruns
                cur_reruns = cur_reruns.sort_values("index")
                def _match(df, value):
                    # df["column"] == None will always be False, so we need special consideration for it.
                    # See: https://pandas.pydata.org/pandas-docs/stable/user_guide/missing_data.html
                    if value is None:
                        return df.isna()
                    return df == value
                non_reruns = df[_match(df[self.config.input.sample_id_col], group_id) & \
                    _match(df[self.config.input.target_col], group_target) & \
                    _match(df[self.config.input.measure_type_col], group_measure_type)]
                non_reruns = non_reruns.sort_values(self.config.input.index_col)
                df = df.drop(index=non_reruns.index[:min(len(cur_reruns.index), len(non_reruns.index))])

                # Rename the reruns to be a non-rerun
                df.loc[cur_reruns.index, self.config.input.sample_id_col] = group_id

        return df
        
    def get_standard_curve_common_target(self, target):
        """Get the common root target name that is used for the specified target to refer to a standard curve. For example,
        the target PMMoV:10, which is a PMMoV dilution, will typically be mapped to the target PMMoV, which is
        the undiluted form. Both PMMoV:10 and PMMoV use the same standard curve.
        """
        def _get_common_target(target):
            common_targets = self.config.input.get("standard_curve_common_targets", None)
            if common_targets is not None:
                for key, names in common_targets.items():
                    if target.lower() in [n.lower() for n in names]:
                        return key
            return target
        if isinstance(target, str):
            target = _get_common_target(target)
        else:
            target = target.map(_get_common_target)
        return target

    def get_standard_curve_id(self, target, plateID):
        """Get the standard curve ID (ie. it's sheet name) for the standard curve for the specified target
        on the specified plateID. 

        Parameters
        ----------
        target : str | pd.DataFrame
            The target to get the curve ID for. Can be a single target (str) or a DataFrame of targets. If a DataFrame
            then the column self.config.input.target_col is used.
        plateID : str | pd.DataFrame
            The plate ID that contains the standard curve for the target. Can be a single plate ID (str) or a DataFrame 
            of plate IDs. If a DataFrame then the column self.config.input.plate_id_col is used.

        Returns
        -------
        str | pd.Series
            The standard curve ID(s). If both target and plateID are strings then a single string is returned, otherwise
            a pd.Series is returned, with all curve IDs for the input target/plateID with matching index.
        """
        target = self.get_standard_curve_common_target(target)

        if isinstance(target, str) and isinstance(plateID, str):
            return CAL_SHEET_FMT.format(targetName=target, plateID=plateID)
        elif isinstance(target, str):
            # target is a string, plateID is a DataFrame
            target = pd.DataFrame({self.config.input.target_col : [target]*len(plateID.index)}, index=plateID.index)
        elif isinstance(plateID, str):
            # target is a DataFrame, plateID is a string
            plateID = pd.DataFrame({self.config.input.plate_id_col : [plateID]*len(target.index)}, index=target.index)

        return pd.concat([target, plateID], axis=1).agg(lambda x: CAL_SHEET_FMT.format(targetName=x[self.config.input.target_col], plateID=x[self.config.input.plate_id_col]), axis=1)

        # return CAL_SHEET_FMT.format(targetName=target, plateID=plateID)

    def assign_standard_curve_ids(self):
        """Add the standard curve IDs to all items in the QPCR DataFrame (self.qpcr_df). The standard curve IDs depend
        on the target and plate ID.
        """
        self.qpcr_df[self.config.input.standard_curve_id_col] = None

        # Assign all curve plate IDs for all std samples. This will give us all standardCurveIDs that are available
        # in the QPCR data.
        std_filt = self.qpcr_df[self.config.input.measure_type_col] == self.config.input.measure_type_std
        self.qpcr_df.loc[std_filt, self.config.input.standard_curve_id_col] = self.get_standard_curve_id(self.qpcr_df.loc[std_filt, self.config.input.target_col], self.qpcr_df.loc[std_filt, self.config.input.plate_id_col])
        
        known_std_curves = self.qpcr_df.loc[std_filt]
        known_std_curves = known_std_curves.sort_values(self.config.input.analysis_date_col, ascending=False)

        def _find_standard_curve(row):
            same_target_filt = self.get_standard_curve_common_target(row[self.config.input.target_col]) == self.get_standard_curve_common_target(known_std_curves[self.config.input.target_col])
            same_plate_id_filt = row[self.config.input.plate_id_col] == known_std_curves[self.config.input.plate_id_col] #standard_curve_id_col]

            # Our own plate has a standard curve for our target type
            if (same_target_filt & same_plate_id_filt).sum() > 0:
                return self.get_standard_curve_id(row[self.config.input.target_col], row[self.config.input.plate_id_col])

            if self.config.input.require_cal_curve_on_same_plate:
                # Standard curves must be on the same plate as the unknowns, but none was found
                return ""

            # Try to find another plate with a standard curve for our target type
            # We take the plate with the QPCR date closest to our own QPCR date (either before or after)
            other_plates_filt = known_std_curves[self.config.input.plate_id_col] != row[self.config.input.plate_id_col]
            other_plates = known_std_curves[other_plates_filt & same_target_filt]
            if len(other_plates.index) == 0:
                # No standard curves with same target (but on different plate) found
                return ""

            delta_date = (other_plates[self.config.input.analysis_date_col] - row[self.config.input.analysis_date_col]).map(abs)
            min_delta = delta_date.idxmin()
            return other_plates.loc[min_delta][self.config.input.standard_curve_id_col]

        # Assign all curve plate IDs for all QPCR samples
        ct_filt = self.qpcr_df[self.config.input.measure_type_col].isin([self.config.input.measure_type_unknown, self.config.input.measure_type_ntc, self.config.input.measure_type_eb])
        self.qpcr_df.loc[ct_filt, self.config.input.standard_curve_id_col] = self.qpcr_df.loc[ct_filt].agg(_find_standard_curve, axis=1)

        # Remove all Unknowns that do not have a standard curve
        # if self.config.input.require_cal_curve_on_same_plate:
        unknowns_filt = self.qpcr_df[self.config.input.measure_type_col] == self.config.input.measure_type_unknown
        has_curve_filt = self.qpcr_df[self.config.input.standard_curve_id_col] != ""
        self.qpcr_df = self.qpcr_df[~unknowns_filt | (unknowns_filt & has_curve_filt)]
              
    def apply_value_mappers(self):
        """Perform additional processing by running the value_mappers in the config file on the main QPCR input. eg. samples named "EB" will
        receive "EB" in the "Content" column.
        """
        if "value_mappers" in self.config.input:
            for value_mapper in self.config.input.value_mappers:
                filt = self.qpcr_df[value_mapper.match_column].str.contains(value_mapper.match_expression, case=value_mapper.get("ignore_case", False))
                filt = filt.fillna(False)
                self.qpcr_df.loc[filt, value_mapper.target_column] = value_mapper.target_value
                                        
    def populate(self):
        """Do a full population of the output. This is the main function to call by a QPCRPopulator user.
        """        
        output_files = []
        # target_dir = os.path.dirname(self.local_target_file)
        # if target_dir and not os.path.exists(target_dir):
        #     os.makedirs(target_dir, exist_ok=True)
        self.template_xl = openpyxl.load_workbook(self.template_file)
        print(f"Loading input file {self.input_file}...")
        fix_xlsx_file(self.input_file)
        # input_xl = openpyxl.load_workbook(self.input_file)
        input_df = pd.read_excel(self.input_file, sheet_name=0)
                
        print(f"Loading QPCR sheet {self.input_file}...")
        # self.measure_sheet_df = sheet_to_df(input_xl[self.config.input.measure_sheet_name])
        self.qpcr_df = input_df #sheet_to_df(input_xl[input_xl.sheetnames[0]])
        self.apply_value_mappers()
        self.qpcr_df = self.sampleids.make_all_sample_ids(
            self.qpcr_df, 
            sample_id_col=self.config.input.sample_id_col, 
            sample_date_col=None, 
            target_sample_id_col=self.config.input.sample_id_col,
            target_match_sample_id_col=self.config.input.match_sample_id_col)
                
        # print(f"Loading sample sheet {self.config.input.sample_sheet_name}...")
        # sample_sheet_df = sheet_to_df(input_xl[self.config.input.sample_sheet_name])
        # sample_sheet_df[self.config.input.sample_short_description_col] = self.sites.get_type_short_description(sample_sheet_df[self.config.input.sample_type_col]) if self.sites is not None else ""
        # sample_sheet_df[self.config.input.sample_description_col] = self.sites.get_type_description(sample_sheet_df[self.config.input.sample_type_col]) if self.sites is not None else ""

        # Add samples log data to QPCR sheet
        self.qpcr_df = self.sampleslog.join_and_cast(self.qpcr_df, on=self.config.input.match_sample_id_col)
        
        # Populate site ID, based on the sample ID
        self.qpcr_df[self.config.input.site_id_col] = self.qpcr_df[self.config.input.sample_id_col].apply(self.sites.get_siteid_from_sampleid)

        # Add sample type info (eg. pSludge, pEfflu, water, etc)
        self.qpcr_df[self.config.input.sampling_type_col] = self.sites.get_site_sample_type(self.qpcr_df[self.config.input.site_id_col])
        self.qpcr_df[self.config.input.sample_short_description_col] = self.sites.get_type_short_description(self.qpcr_df[self.config.input.sampling_type_col]) if self.sites is not None else ""
        self.qpcr_df[self.config.input.sample_description_col] = self.sites.get_type_description(self.qpcr_df[self.config.input.sampling_type_col]) if self.sites is not None else ""

        self.assign_standard_curve_ids()
        
        # Create (overwrite) sample IDs for all std entries.
        std_filt = self.qpcr_df[self.config.input.measure_type_col] == self.config.input.measure_type_std
        self.qpcr_df.loc[std_filt, self.config.input.sample_id_col] = self.qpcr_df[std_filt][[self.config.input.sq_col, self.config.input.plate_id_col]].agg(lambda x: f"{x[1]}-{x[0]}", axis=1)

        # Convert target names to recognized ones
        self.qpcr_df[self.config.input.target_col] = self.qpcr_df[self.config.input.target_col].map(self.get_recognized_target)
        
        #  Set the index. This allows us to preserve the original ordering if needed
        for group_name, group_df in self.qpcr_df.groupby([self.config.input.target_col, self.config.input.sample_id_col, self.config.input.measure_type_col]):
            self.qpcr_df.loc[group_df.index, self.config.input.index_col] = list(np.arange(0, len(group_df.index)))
        
        self.qpcr_df = self.qpcr_df.sort_values(self.config.input.order_by)

        # Prepare reruns.
        self.qpcr_df = self.prepare_reruns(self.qpcr_df)

        # Create blank outliers column. Will populate outliers later on
        self.qpcr_df[OUTLIER_COL] = None

        # Split the QPCR data into groups. Each group represents a single output file.
        for file_info, file_group_df in self.make_file_splits(self.qpcr_df):
            target_file = file_info["fileName"]
            local_target_file = cloud_utils.download_file(target_file)
            target_dir = os.path.dirname(local_target_file) if local_target_file else ""
            if target_dir and not os.path.exists(target_dir):
                os.makedirs(target_dir, exist_ok=True)            

            self.worksheets_info = []
            self.late_binders = []

            origin = self.config.template.main_origin.copy()
            extents = origin
            main_ws = None

            # If appending to the output (rather than overwriting), then load the existing output file and update the origin and extents
            if not self.overwrite:
                try:
                    fix_xlsx_file(local_target_file)
                    self.output_wb = openpyxl.load_workbook(local_target_file)
                    if MAIN_SHEET in self.output_wb.sheetnames:
                        main_ws = self.output_wb[MAIN_SHEET]                
                    else:
                        main_ws = self.output_wb[0]
                    origin = [main_ws.max_row + (self.config.template.rows_between_main_groups + 1 if main_ws.max_row > 1 else 0), origin[1]]
                    extents = origin
                except:
                    print("Could not append to output file, creating from scratch.")

            # If no main worksheet then create a new one
            if main_ws is None:
                self.output_wb = Workbook()
                self.output_wb.loaded_theme = self.template_xl.loaded_theme
                for style_name in set(self.template_xl.style_names).difference(self.output_wb.style_names):
                    style = [s for s in self.template_xl._named_styles if s.name == style_name][0]
                    self.output_wb.add_named_style(copy(style))
                main_ws = self.output_wb.active

            main_ws.title = MAIN_SHEET
            self.worksheets_info.append({
                "sheet_name" : main_ws.title,
                "ws" : main_ws,
                "origin" : origin,
                "extents" : extents,
            })

            # Create a separate group for each analysis group (ie. for each analysis day). We will loop through each group
            # and output the results in order.
            analysis_groups = self.make_inner_splits(file_group_df)
            main_has_data = False
            total_groups = len(analysis_groups)
            for idx, (name, group) in enumerate(analysis_groups):
                print(f"Creating group {name} ({idx+1}/{total_groups})...")

                if self.qaqc is not None:
                    self.qaqc.set_current_name(name)

                group = group.copy()
                # self.remove_all_outliers(group)
                self.qaqc.remove_all_outliers(group, self.qpcr_df)
                if not self.create_main(name, group):
                    continue
                main_has_data = True
                self.create_calibration(group)
                self.consolidate_extents()
                self.handle_late_binders(inner=True)
                self.handle_late_binders(inner=False)
                if not self.hide_qaqc:
                    self.qaqc.run_qaqc(self.output_wb, group, self.qpcr_df)
                    self.qaqc.add_qaqc_to_workbook(self.output_wb)
                self.remove_non_main_sheets_info()
                
                # Advance to where we'll be outputting the next group. We advance a certain number of rows
                # with a new origin at that row.
                _, info = self.get_worksheet_and_info(MAIN_SHEET)
                extents = info["extents"]
                info["extents"] = [extents[0]+self.config.template.rows_between_main_groups, extents[1]]
                origin = info["origin"]
                origin[0] = info["extents"][0]

            # self.handle_late_binders(inner=False)

            if main_has_data:
                # We have data in the main sheet, so calculate all the values of all formulas and save to disk
                # self.remove_non_main_sheets_info()
                output_files.append(local_target_file)
                print("Resaving prior to calculating Excel formulas...")
                # @TODO: If we don't save and reload then most cells do not seem to calculate properly
                # due to various exceptions thrown within PyCel. There should be a better way to fix this.
                self.output_wb.save(local_target_file)
                self.output_wb = openpyxl.load_workbook(local_target_file)
                print("Calculating Excel formulas...")
                add_excel_calculated_values(self.output_wb)
                print(f"Saving to {local_target_file}...")
                self.output_wb.save(local_target_file)

                # If local_target_file != target_file, then it's a remote file, so upload it
                if local_target_file != target_file:
                    cloud_utils.upload_file(local_target_file, target_file)
            else:
                print(f"No data, not saving {local_target_file}")
        
        return output_files

if __name__ == "__main__":
    if "get_ipython" in globals():
        opts = EasyDict({
            # B117
            # "input_file" : "/Users/martinwellman/Documents/Health/Wastewater/Code/odmdata/odm_merged-jan6.xlsx",
            # "template_file" : "qpcr_template_b117.xlsx",
            # "config" : ["qpcr_populator_long-2main-inh.yaml", "qpcr_populator_b117_diff.yaml"],
            # "qaqc" : "qaqc_b117.yaml",

            "input_file" : "/Users/martinwellman/Documents/Health/Wastewater/Code/output/extracted-merged.xlsx",

            # Long format
            "template_file" : "qpcr_template_long-2main-inh.xlsx",
            "config" : ["qpcr_populator_long-2main-inh.yaml"],
            "qaqc" : ["qaqc_long-2main-inh.yaml"],

            # Wide format
            # "template_file" : "qpcr_template_wide-2main-inh.xlsx",
            # "config" : ["qpcr_populator_long-2main-inh.yaml", "qpcr_populator_wide_diff-2main-inh.yaml"],
            # "qaqc" : ["qaqc_long-2main-inh.yaml", "qaqc_wide_diff-2main-inh.yaml"],

            "target_file" : "/Users/martinwellman/Documents/Health/Wastewater/Code/output/output.xlsx",
            "hide_qaqc" : False,
            "overwrite" : True,

            "sites_config" : "qpcr_sites.yaml",
            "sites_file" : "qpcr_sites.xlsx",
            
            "sampleids_config" : "qpcr_sampleids.yaml",
            "sampleslog_file" : "/Users/martinwellman/Documents/Health/Wastewater/Code/COVID SAMPLE LOG SHEET-2-7.xlsx",
            "sampleslog_config" : "qpcr_sampleslog.yaml",
            
            "methods_config" : "qpcr_methods.yaml",
            "methods_file" : "/Users/martinwellman/Documents/Health/Wastewater/Code/odm-qpcr-analyzer/qpcr_analyzer/qpcr_methods.xlsx",
        })
    else:
        args = argparse.ArgumentParser(formatter_class=argparse.ArgumentDefaultsHelpFormatter)
        args.add_argument("--input_file", type=str, help="Input file to analyze", required=True)
        args.add_argument("--qaqc", nargs="+", type=str, help="Configuration file for QAQC. Skip QAQC if not specified.", required=False)
        args.add_argument("--template_file", type=str, help="Template Excel file for format of rows added to the final output", required=True)
        args.add_argument("--target_file", type=str, help="Target Excel file to save the final, formatted output to.", required=True)
        args.add_argument("--config", nargs="+", type=str, help="Configuration file", default="qpcr_populator_long-2main-inh.yaml")
        args.add_argument("--overwrite", help="If set then overwrite the target_file (instead of appending to it).", action="store_true")
        args.add_argument("--hide_qaqc", help="If set then do not show QAQC highlighting or sheets. Outliers will still be removed and marked with square brackets.", action="store_true")
        args.add_argument("--sites_config", type=str, help="Config file for the sites file.", required=True)
        args.add_argument("--sites_file", type=str, help="Excel file specifying all WW sites with information about each site.", required=True)
        args.add_argument("--sampleids_config", type=str, help="Config file for the sample IDs.", required=True)
        args.add_argument("--sampleslog_config", type=str, help="Config file for the samples log file.", required=True)
        args.add_argument("--sampleslog_file", type=str, help="Log file for the samples, containing sample mass, analysis date, etc.", required=True)
        args.add_argument("--methods_config", type=str, help="Config file for the methods.", required=True)
        args.add_argument("--methods_file", type=str, help="Excel file with all the method definitions", required=True)
        
        opts = args.parse_args()

    tic = datetime.now()
    print("Starting populator at:", tic)
    qpcr = QPCRPopulator(
        opts.input_file, 
        opts.template_file, 
        opts.target_file, 
        opts.overwrite, 
        opts.config, 
        opts.qaqc, 
        sites_config=opts.sites_config, 
        sites_file=opts.sites_file, 
        sampleids_config=opts.sampleids_config,
        sampleslog_config=opts.sampleslog_config,
        sampleslog_file=opts.sampleslog_file,
        methods_config=opts.methods_config,
        methods_file=opts.methods_file,
        hide_qaqc=opts.hide_qaqc
        )
    qpcr.populate()
    toc = datetime.now()
    print("Started at:", tic)
    print("Ended at:", toc)
    print("Total duration:", toc - tic)
